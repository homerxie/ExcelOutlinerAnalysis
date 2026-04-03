from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from .models import LoadedTable, MeasurementRecord, TemplateConfig


def load_table(
    input_path: str | Path,
    template: TemplateConfig | None = None,
    header_row: int = 0,
) -> LoadedTable:
    path = Path(input_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        measurement_header_row = header_row
        row_header_row = header_row
        unit_row: int | None = None
        data_start_row: int | None = None
        if template is not None:
            measurement_header_row = (
                template.measurement_header_row
                if template.measurement_header_row is not None
                else template.header_row
            )
            row_header_row = (
                template.row_header_row
                if template.row_header_row is not None
                else measurement_header_row
            )
            unit_row = template.unit_row
            data_start_row = template.data_start_row
        return _load_xlsx(
            path,
            measurement_header_row=measurement_header_row,
            row_header_row=row_header_row,
            unit_row=unit_row,
            data_start_row=data_start_row,
        )
    raise ValueError(f"Unsupported file type: {suffix}")


def load_rows(
    input_path: str | Path,
    header_row: int = 0,
    template: TemplateConfig | None = None,
) -> list[dict[str, Any]]:
    return load_table(input_path, template=template, header_row=header_row).rows


def build_measurements(
    template: TemplateConfig,
    rows: list[dict[str, Any]],
    dataset_id: str,
    source_file: str,
) -> list[MeasurementRecord]:
    measurements: list[MeasurementRecord] = []
    for row_index, row in enumerate(rows, start=1):
        dimensions = extract_dimensions(template, row)
        if dimensions is None:
            continue
        for group in template.analysis_groups:
            for column in group.columns:
                if column not in row:
                    continue
                value = to_float(row[column])
                if value is None:
                    continue
                logical_metric = (
                    group.id
                    if group.analysis_mode == "pooled_columns"
                    else f"{group.id}::{column}"
                )
                measurements.append(
                    MeasurementRecord(
                        dataset_id=dataset_id,
                        source_file=source_file,
                        row_number=row_index,
                        logical_metric=logical_metric,
                        group_id=group.id,
                        group_display_name=group.display_name,
                        presentation_group=group.presentation_group,
                        raw_column=column,
                        value=value,
                        dimensions=dimensions,
                    )
                )
    return measurements


def extract_dimensions(
    template: TemplateConfig, row: dict[str, Any]
) -> dict[str, str] | None:
    dimensions: dict[str, str] = {}
    for dimension in template.row_dimensions:
        parts: list[str] = []
        for source in dimension.sources:
            raw_value = row.get(source.column)
            if raw_value is None or str(raw_value).strip() == "":
                continue
            text = str(raw_value).strip()
            if source.split_position is None:
                parts.append(text)
                continue
            split_parts = _split_by_delimiters(text, source.split_delimiters)
            if source.skip_empty_parts:
                split_parts = [item for item in split_parts if item != ""]
            if source.split_position >= len(split_parts):
                continue
            part = split_parts[source.split_position].strip()
            if part:
                parts.append(part)
        if not parts:
            if dimension.default_value is not None:
                dimensions[dimension.name] = dimension.default_value
                continue
            if dimension.optional:
                continue
            return None
        dimensions[dimension.name] = "|".join(parts)
    return dimensions


def _split_by_delimiters(value: str, delimiters: list[str]) -> list[str]:
    if not delimiters:
        return [value]
    pattern = "|".join(re.escape(item) for item in sorted(delimiters, key=len, reverse=True))
    if not pattern:
        return [value]
    return re.split(pattern, value)


def _load_csv(path: Path) -> LoadedTable:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = [dict(row) for row in reader]
        return LoadedTable(headers=list(reader.fieldnames or []), rows=rows)


def _load_xlsx(
    path: Path,
    measurement_header_row: int = 0,
    row_header_row: int = 0,
    unit_row: int | None = None,
    data_start_row: int | None = None,
) -> LoadedTable:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading xlsx requires openpyxl. Run `pip3 install -e .` first."
        ) from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return LoadedTable(headers=[], rows=[], units={})

    measurement_headers = _normalize_header_row(rows, measurement_header_row)
    row_headers = _normalize_header_row(rows, row_header_row)
    units = (
        _normalize_header_row(rows, unit_row)
        if unit_row is not None
        else [""] * max(len(measurement_headers), len(row_headers))
    )
    headers = _combine_headers(measurement_headers, row_headers)
    if data_start_row is None:
        candidate_rows = [measurement_header_row, row_header_row]
        if unit_row is not None:
            candidate_rows.append(unit_row)
        data_start_row = max(candidate_rows) + 1
    data_rows = rows[data_start_row:]
    normalized_rows: list[dict[str, Any]] = []
    for values in data_rows:
        payload = {
            headers[index]: values[index] if index < len(values) else None
            for index in range(len(headers))
            if headers[index]
        }
        if any(value is not None and str(value).strip() != "" for value in payload.values()):
            normalized_rows.append(payload)
    unit_map = {
        headers[index]: units[index]
        for index in range(len(headers))
        if headers[index] and index < len(units) and units[index]
    }
    return LoadedTable(headers=headers, rows=normalized_rows, units=unit_map)


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_header_row(rows: list[tuple[Any, ...]], row_index: int | None) -> list[str]:
    if row_index is None or row_index >= len(rows):
        return []
    return [str(item).strip() if item is not None else "" for item in rows[row_index]]


def _combine_headers(measurement_headers: list[str], row_headers: list[str]) -> list[str]:
    size = max(len(measurement_headers), len(row_headers))
    headers: list[str] = []
    used: dict[str, int] = {}
    for index in range(size):
        header = ""
        if index < len(row_headers) and row_headers[index]:
            header = row_headers[index]
        elif index < len(measurement_headers) and measurement_headers[index]:
            header = measurement_headers[index]
        if header and header in used:
            used[header] += 1
            header = f"{header}__{used[header]}"
        elif header:
            used[header] = 1
        headers.append(header)
    return headers
