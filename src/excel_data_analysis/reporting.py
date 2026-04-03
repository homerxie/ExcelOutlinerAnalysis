from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .analyzer import evaluate_against_golden, evaluate_modified_zscore
from .io import extract_dimensions, load_table, to_float
from .models import (
    AnomalyResult,
    GoldenReference,
    MeasurementRecord,
    OutlierRatioStatConfig,
    TemplateConfig,
    ThresholdConfig,
)
from .output_paths import resolve_output_path
from .template import load_template


HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF2F8")
FAIL_FONT = Font(color="9C0006", bold=True)
HEADER_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center")
PERCENT_FORMAT_2 = "0.00%"
DECIMAL_FORMAT_3 = "0.000"


@dataclass(slots=True)
class ColumnMeta:
    raw_column: str
    logical_metric: str
    group_id: str
    chain_name: str
    item_label: str
    repeat_index: int
    unit: str | None
    display_order: int
    metric_type: str
    condition: str


@dataclass(slots=True)
class RowData:
    dataset_id: str
    source_file: str
    row_number: int
    dimensions: dict[str, str]
    values: dict[str, float]


@dataclass(slots=True)
class MetricValue:
    value: float | None
    threshold: float | None
    golden_value: float | None
    is_fail: bool


MetricKey = tuple[str, int, str]


def _measurement_key(measurement: MeasurementRecord) -> MetricKey:
    return (measurement.dataset_id, measurement.row_number, measurement.raw_column)


def _anomaly_key(anomaly: AnomalyResult) -> MetricKey:
    return (anomaly.dataset_id, anomaly.row_number, anomaly.raw_column)


def _row_metric_key(row: RowData, raw_column: str) -> MetricKey:
    return (row.dataset_id, row.row_number, raw_column)


def generate_chip_report(
    template_path: str | Path,
    input_path: str | Path,
    output_path: str | Path,
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
    if_exists: str = "overwrite",
) -> dict[str, Any]:
    resolved_output_path = resolve_output_path(output_path, if_exists=if_exists)
    resolved_output_path.parent.mkdir(parents=True, exist_ok=True)
    template = load_template(template_path)
    table = load_table(input_path, template=template)
    measurements = _build_measurements(template, table.rows, input_path)
    row_data = _build_rows(template, table.rows, input_path)
    columns = _build_columns(template, table.units, table.headers)
    return _generate_chip_report_from_prebuilt(
        template=template,
        source_label=str(Path(input_path).resolve()),
        output_path=resolved_output_path,
        measurements=measurements,
        row_data=row_data,
        columns=columns,
        built_golden_path=built_golden_path,
        outlier_fail_method_override=outlier_fail_method_override,
        zscore_threshold_override=zscore_threshold_override,
    )


def generate_chip_report_from_measurements(
    template_path: str | Path,
    measurements: list[MeasurementRecord],
    output_path: str | Path,
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
    source_label: str | None = None,
    if_exists: str = "overwrite",
) -> dict[str, Any]:
    resolved_output_path = resolve_output_path(output_path, if_exists=if_exists)
    resolved_output_path.parent.mkdir(parents=True, exist_ok=True)
    template = load_template(template_path)
    row_data = _build_rows_from_measurements(measurements)
    columns = _build_columns(template, {}, [])
    resolved_source_label = source_label or "database-scope"
    return _generate_chip_report_from_prebuilt(
        template=template,
        source_label=resolved_source_label,
        output_path=resolved_output_path,
        measurements=measurements,
        row_data=row_data,
        columns=columns,
        built_golden_path=built_golden_path,
        outlier_fail_method_override=outlier_fail_method_override,
        zscore_threshold_override=zscore_threshold_override,
    )


def _generate_chip_report_from_prebuilt(
    template: TemplateConfig,
    source_label: str,
    output_path: str | Path,
    measurements: list[MeasurementRecord],
    row_data: list[RowData],
    columns: list[ColumnMeta],
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
) -> dict[str, Any]:
    sample_node_orders = _build_sample_node_orders(row_data, template)
    row_data.sort(key=lambda item: _row_sort_key(item, template, sample_node_orders))

    zscore_values = _build_zscore_map(
        measurements,
        template.report.zscore_thresholds,
        columns,
        zscore_threshold_override=zscore_threshold_override,
    )
    golden_values = _build_report_golden_values(
        measurements,
        template,
        columns,
        built_golden_path=built_golden_path,
    )
    golden_header_values = _build_column_display_values(golden_values, "golden_value")
    golden_threshold_values = _build_column_display_values(golden_values, "threshold")
    zscore_threshold_values = _build_column_display_values(zscore_values, "threshold")

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sorted_sheet(
        workbook.create_sheet("Sorted"),
        template,
        columns,
        row_data,
        golden_header_values,
    )
    if built_golden_path and _has_variable_metric_attribute(golden_values, "golden_value"):
        _write_point_golden_sheet(
            workbook.create_sheet("golden-by-point"),
            template,
            columns,
            row_data,
            golden_values,
            golden_header_values,
        )
    _write_metric_sheet(
        workbook.create_sheet("deviation-zscore"),
        template,
        columns,
        row_data,
        zscore_values,
        zscore_threshold_values,
        golden_header_values,
        "Z Threshold",
        value_number_format=DECIMAL_FORMAT_3,
        threshold_number_format=DECIMAL_FORMAT_3,
    )
    _write_metric_sheet(
        workbook.create_sheet("deviation-golden"),
        template,
        columns,
        row_data,
        golden_values,
        golden_threshold_values,
        golden_header_values,
        "Deviation Threshold",
        value_number_format=PERCENT_FORMAT_2,
        threshold_number_format=PERCENT_FORMAT_2,
    )
    summary_artifacts = _collect_outlier_summary_artifacts(
        template,
        columns,
        row_data,
        zscore_values,
        golden_values,
        sample_node_orders,
        outlier_fail_method_override=outlier_fail_method_override,
    )
    summary_rows = summary_artifacts["summary_rows"]
    ratio_rows = summary_artifacts["ratio_rows"]
    _write_outlier_sheet(
        workbook.create_sheet("outliners"),
        template,
        columns,
        row_data,
        zscore_values,
        golden_values,
        golden_header_values,
        sample_node_orders,
        summary_rows,
        outlier_fail_method_override=outlier_fail_method_override,
    )
    _write_summary_sheet(
        workbook.create_sheet("Outliner Summary"),
        template,
        summary_rows,
        ratio_rows,
        sample_node_orders,
    )
    workbook.save(output_path)

    failing_samples = sorted({item["sample_id"] for item in summary_rows})
    return {
        "input": source_label,
        "output": str(Path(output_path).resolve()),
        "row_count": len(row_data),
        "measurement_count": len(measurements),
        "outlier_sample_count": len(failing_samples),
        "outlier_event_count": len(summary_rows),
    }


def collect_report_failures(
    template_path: str | Path,
    input_path: str | Path,
    built_golden_path: str | Path | None = None,
    zscore_threshold_override: float | None = None,
) -> list[AnomalyResult]:
    template = load_template(template_path)
    table = load_table(input_path, template=template)
    measurements = _build_measurements(template, table.rows, input_path)
    columns = _build_columns(template, table.units, table.headers)
    zscore_values = _build_zscore_map(
        measurements,
        template.report.zscore_thresholds,
        columns,
        zscore_threshold_override=zscore_threshold_override,
    )
    golden_values = _build_report_golden_values(
        measurements,
        template,
        columns,
        built_golden_path=built_golden_path,
    )

    failures: list[AnomalyResult] = []
    for measurement in measurements:
        key = _measurement_key(measurement)
        z_entry = zscore_values.get(key)
        if z_entry and z_entry.is_fail:
            failures.append(
                AnomalyResult(
                    method="modified_z_score",
                    dataset_id=measurement.dataset_id,
                    source_file=measurement.source_file,
                    row_number=measurement.row_number,
                    logical_metric=measurement.logical_metric,
                    raw_column=measurement.raw_column,
                    value=measurement.value,
                    dimensions=measurement.dimensions,
                    center=None,
                    score=z_entry.value,
                    reason=(
                        "Absolute modified z-score exceeded configured threshold "
                        f"{z_entry.threshold}."
                    ),
                )
            )
        golden_entry = golden_values.get(key)
        if golden_entry and golden_entry.is_fail:
            failures.append(
                AnomalyResult(
                    method="golden_deviation",
                    dataset_id=measurement.dataset_id,
                    source_file=measurement.source_file,
                    row_number=measurement.row_number,
                    logical_metric=measurement.logical_metric,
                    raw_column=measurement.raw_column,
                    value=measurement.value,
                    dimensions=measurement.dimensions,
                    center=golden_entry.golden_value,
                    score=golden_entry.value,
                    reason=(
                        "Golden deviation exceeded configured threshold "
                        f"{golden_entry.threshold}."
                    ),
                )
            )
    return failures


def collect_outlier_summary_rows(
    template_path: str | Path,
    input_path: str | Path,
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
) -> list[dict[str, str]]:
    return collect_outlier_summary_artifacts(
        template_path,
        input_path,
        built_golden_path=built_golden_path,
        outlier_fail_method_override=outlier_fail_method_override,
        zscore_threshold_override=zscore_threshold_override,
    )["summary_rows"]


def collect_outlier_ratio_rows(
    template_path: str | Path,
    input_path: str | Path,
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
) -> list[dict[str, Any]]:
    return collect_outlier_summary_artifacts(
        template_path,
        input_path,
        built_golden_path=built_golden_path,
        outlier_fail_method_override=outlier_fail_method_override,
        zscore_threshold_override=zscore_threshold_override,
    )["ratio_rows"]


def collect_outlier_summary_artifacts(
    template_path: str | Path,
    input_path: str | Path,
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
) -> dict[str, list[dict[str, Any]]]:
    template = load_template(template_path)
    table = load_table(input_path, template=template)
    measurements = _build_measurements(template, table.rows, input_path)
    row_data = _build_rows(template, table.rows, input_path)
    columns = _build_columns(template, table.units, table.headers)
    return _collect_outlier_summary_artifacts_from_prebuilt(
        template,
        measurements,
        row_data,
        columns,
        built_golden_path=built_golden_path,
        outlier_fail_method_override=outlier_fail_method_override,
        zscore_threshold_override=zscore_threshold_override,
    )


def collect_outlier_summary_rows_from_measurements(
    template_path: str | Path,
    measurements: list[MeasurementRecord],
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
) -> list[dict[str, str]]:
    return collect_outlier_summary_artifacts_from_measurements(
        template_path,
        measurements,
        built_golden_path=built_golden_path,
        outlier_fail_method_override=outlier_fail_method_override,
        zscore_threshold_override=zscore_threshold_override,
    )["summary_rows"]


def collect_outlier_ratio_rows_from_measurements(
    template_path: str | Path,
    measurements: list[MeasurementRecord],
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
) -> list[dict[str, Any]]:
    return collect_outlier_summary_artifacts_from_measurements(
        template_path,
        measurements,
        built_golden_path=built_golden_path,
        outlier_fail_method_override=outlier_fail_method_override,
        zscore_threshold_override=zscore_threshold_override,
    )["ratio_rows"]


def collect_outlier_summary_artifacts_from_measurements(
    template_path: str | Path,
    measurements: list[MeasurementRecord],
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
) -> dict[str, list[dict[str, Any]]]:
    template = load_template(template_path)
    row_data = _build_rows_from_measurements(measurements)
    columns = _build_columns(template, {}, [])
    return _collect_outlier_summary_artifacts_from_prebuilt(
        template,
        measurements,
        row_data,
        columns,
        built_golden_path=built_golden_path,
        outlier_fail_method_override=outlier_fail_method_override,
        zscore_threshold_override=zscore_threshold_override,
    )


def _collect_outlier_summary_rows_from_prebuilt(
    template: TemplateConfig,
    measurements: list[MeasurementRecord],
    row_data: list[RowData],
    columns: list[ColumnMeta],
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
) -> list[dict[str, str]]:
    return _collect_outlier_summary_artifacts_from_prebuilt(
        template,
        measurements,
        row_data,
        columns,
        built_golden_path=built_golden_path,
        outlier_fail_method_override=outlier_fail_method_override,
        zscore_threshold_override=zscore_threshold_override,
    )["summary_rows"]


def _collect_outlier_summary_artifacts_from_prebuilt(
    template: TemplateConfig,
    measurements: list[MeasurementRecord],
    row_data: list[RowData],
    columns: list[ColumnMeta],
    built_golden_path: str | Path | None = None,
    outlier_fail_method_override: str | None = None,
    zscore_threshold_override: float | None = None,
) -> dict[str, list[dict[str, Any]]]:
    sample_node_orders = _build_sample_node_orders(row_data, template)
    row_data.sort(key=lambda item: _row_sort_key(item, template, sample_node_orders))
    zscore_values = _build_zscore_map(
        measurements,
        template.report.zscore_thresholds,
        columns,
        zscore_threshold_override=zscore_threshold_override,
    )
    golden_values = _build_report_golden_values(
        measurements,
        template,
        columns,
        built_golden_path=built_golden_path,
    )
    return _collect_outlier_summary_artifacts(
        template,
        columns,
        row_data,
        zscore_values,
        golden_values,
        sample_node_orders,
        outlier_fail_method_override=outlier_fail_method_override,
    )


def _build_measurements(
    template: TemplateConfig,
    rows: list[dict[str, Any]],
    input_path: str | Path,
) -> list[MeasurementRecord]:
    source_file = str(Path(input_path).resolve())
    measurements: list[MeasurementRecord] = []
    for row_number, row in enumerate(rows, start=1):
        dimensions = extract_dimensions(template, row)
        if dimensions is None:
            continue
        for group in template.analysis_groups:
            for column in group.columns:
                if column not in row:
                    continue
                value = to_float(row.get(column))
                if value is None:
                    continue
                logical_metric = (
                    group.id
                    if group.analysis_mode == "pooled_columns"
                    else f"{group.id}::{column}"
                )
                measurements.append(
                    MeasurementRecord(
                        dataset_id="ad-hoc-report",
                        source_file=source_file,
                        row_number=row_number,
                        logical_metric=logical_metric,
                        group_id=group.id,
                        group_display_name=group.display_name,
                        presentation_group=group.presentation_group,
                        raw_column=column,
                        value=value,
                        dimensions=dimensions,
                    )
                )
    return measurements


def _build_rows(
    template: TemplateConfig,
    rows: list[dict[str, Any]],
    input_path: str | Path,
) -> list[RowData]:
    report_rows: list[RowData] = []
    source_file = str(Path(input_path).resolve())
    measurement_columns = {
        column
        for group in template.analysis_groups
        for column in group.columns
    }
    for row_number, row in enumerate(rows, start=1):
        dimensions = extract_dimensions(template, row)
        if dimensions is None:
            continue
        values: dict[str, float] = {}
        for column in measurement_columns:
            if column not in row:
                continue
            value = to_float(row.get(column))
            if value is not None:
                values[column] = value
        report_rows.append(
            RowData(
                dataset_id="ad-hoc-report",
                source_file=source_file,
                row_number=row_number,
                dimensions=dimensions,
                values=values,
            )
        )
    return report_rows


def _build_rows_from_measurements(measurements: list[MeasurementRecord]) -> list[RowData]:
    grouped: dict[tuple[str, str, int, tuple[tuple[str, str], ...]], RowData] = {}
    for item in measurements:
        dimension_key = tuple(sorted(item.dimensions.items()))
        row_key = (item.dataset_id, item.source_file, item.row_number, dimension_key)
        row = grouped.get(row_key)
        if row is None:
            row = RowData(
                dataset_id=item.dataset_id,
                source_file=item.source_file,
                row_number=item.row_number,
                dimensions=dict(item.dimensions),
                values={},
            )
            grouped[row_key] = row
        row.values[item.raw_column] = item.value
    return list(grouped.values())


def _build_columns(
    template: TemplateConfig,
    units: dict[str, str],
    headers: list[str],
) -> list[ColumnMeta]:
    header_set = set(headers)
    columns: list[ColumnMeta] = []
    display_order = 0
    chain_order: dict[str, int] = {}
    for group in template.analysis_groups:
        for column in group.columns:
            if header_set and column not in header_set:
                continue
            parsed = _parse_measurement_column(column)
            if parsed["chain_name"] not in chain_order:
                chain_order[parsed["chain_name"]] = len(chain_order)
            logical_metric = (
                group.id
                if group.analysis_mode == "pooled_columns"
                else f"{group.id}::{column}"
            )
            columns.append(
                ColumnMeta(
                    raw_column=column,
                    logical_metric=logical_metric,
                    group_id=group.id,
                    chain_name=parsed["chain_name"],
                    item_label=parsed["item_label"],
                    repeat_index=parsed["repeat_index"],
                    unit=units.get(column) or group.unit,
                    display_order=display_order,
                    metric_type=parsed["metric_type"],
                    condition=parsed["condition"],
                )
            )
            display_order += 1
    return sorted(
        columns,
        key=lambda item: (
            chain_order[item.chain_name],
            _condition_sort_key(item.condition),
            _metric_sort_key(item.metric_type),
            item.repeat_index,
            item.display_order,
        ),
    )


def _build_zscore_map(
    measurements: list[MeasurementRecord],
    thresholds: ThresholdConfig,
    columns: list[ColumnMeta],
    zscore_threshold_override: float | None = None,
) -> dict[MetricKey, MetricValue]:
    column_index = {item.raw_column: item for item in columns}
    evaluations = evaluate_modified_zscore(measurements)
    values: dict[MetricKey, MetricValue] = {}
    for entry in evaluations:
        meta = column_index.get(entry.raw_column)
        if meta is None:
            continue
        threshold = zscore_threshold_override
        if threshold is None:
            threshold = _resolve_threshold(
                thresholds,
                [entry.logical_metric, entry.raw_column, meta.group_id, meta.chain_name],
            )
        score = abs(entry.score) if entry.score is not None else 0.0
        values[_anomaly_key(entry)] = MetricValue(
            value=score,
            threshold=threshold,
            golden_value=None,
            is_fail=threshold is not None and abs(score) > threshold,
        )
    return values


def _build_golden_deviation_map(
    measurements: list[MeasurementRecord],
    template: TemplateConfig,
    columns: list[ColumnMeta],
) -> dict[MetricKey, MetricValue]:
    column_index = {item.raw_column: item for item in columns}
    values: dict[MetricKey, MetricValue] = {}
    for entry in measurements:
        meta = column_index.get(entry.raw_column)
        if meta is None:
            continue
        golden_value = _resolve_golden_value(
            template,
            [entry.logical_metric, entry.raw_column, entry.group_id, meta.chain_name],
        )
        threshold = _resolve_threshold(
            template.report.golden_deviation_thresholds,
            [entry.logical_metric, entry.raw_column, entry.group_id, meta.chain_name],
        )
        deviation: float | None = None
        is_fail = False
        if golden_value is not None:
            if golden_value == 0:
                deviation = entry.value - golden_value
            else:
                deviation = (entry.value - golden_value) / abs(golden_value)
            is_fail = threshold is not None and abs(deviation) > threshold
        values[_measurement_key(entry)] = MetricValue(
            value=deviation,
            threshold=threshold,
            golden_value=golden_value,
            is_fail=is_fail,
        )
    return values


def _build_built_golden_deviation_map(
    measurements: list[MeasurementRecord],
    golden_reference: GoldenReference,
) -> dict[MetricKey, MetricValue]:
    values: dict[MetricKey, MetricValue] = {}
    evaluations = evaluate_against_golden(measurements, golden_reference)
    for entry in evaluations:
        golden_value = entry.center
        deviation: float | None = None
        threshold: float | None = None
        if golden_value is not None:
            if golden_value == 0:
                deviation = entry.value - golden_value
                threshold = _resolve_built_absolute_threshold(entry)
            else:
                deviation = (entry.value - golden_value) / abs(golden_value)
                threshold = _resolve_built_relative_threshold(entry)
        values[_anomaly_key(entry)] = MetricValue(
            value=deviation,
            threshold=threshold,
            golden_value=golden_value,
            is_fail=entry.reason is not None,
        )
    return values


def _build_report_golden_values(
    measurements: list[MeasurementRecord],
    template: TemplateConfig,
    columns: list[ColumnMeta],
    built_golden_path: str | Path | None,
) -> dict[MetricKey, MetricValue]:
    if built_golden_path:
        golden_reference = GoldenReference.from_dict(
            json.loads(Path(built_golden_path).read_text(encoding="utf-8"))
        )
        return _build_built_golden_deviation_map(measurements, golden_reference)
    return _build_golden_deviation_map(measurements, template, columns)


def _write_sorted_sheet(
    worksheet,
    template: TemplateConfig,
    columns: list[ColumnMeta],
    rows: list[RowData],
    golden_header_values: dict[str, Any],
) -> None:
    dimension_labels = [item.display_name or item.name for item in template.row_dimensions]
    label_column = len(dimension_labels)
    data_start_column = label_column + 1
    first_row = 1

    worksheet.cell(first_row, label_column, "ChainName")
    worksheet.cell(first_row + 1, label_column, "TestItem")
    worksheet.cell(first_row + 2, label_column, "RepeatIndex")
    worksheet.cell(first_row + 3, label_column, "Golden")
    for offset, column in enumerate(columns):
        cell_column = data_start_column + offset
        worksheet.cell(first_row, cell_column, column.chain_name)
        worksheet.cell(first_row + 1, cell_column, column.item_label)
        worksheet.cell(first_row + 2, cell_column, column.repeat_index)
        worksheet.cell(first_row + 3, cell_column, golden_header_values.get(column.raw_column))
        worksheet.cell(first_row + 4, cell_column, column.unit)

    for index, label in enumerate(dimension_labels, start=1):
        worksheet.cell(first_row + 5, index, label)

    data_start_row = first_row + 6
    for row_offset, row in enumerate(rows):
        target_row = data_start_row + row_offset
        for index, dimension in enumerate(template.row_dimensions, start=1):
            worksheet.cell(
                target_row,
                index,
                _display_dimension_value(row.dimensions.get(dimension.name)),
            )
        for column_offset, column in enumerate(columns):
            worksheet.cell(target_row, data_start_column + column_offset, row.values.get(column.raw_column))

    _style_sheet(
        worksheet,
        header_rows=list(range(1, 6)),
        dimension_header_row=6,
        start_column=label_column,
        end_column=data_start_column + len(columns) - 1,
        dimension_count=len(dimension_labels),
    )
    _merge_chain_headers(worksheet, first_row, data_start_column, columns)
    _merge_dimension_values(worksheet, data_start_row, len(rows), template)
    _autosize_columns(worksheet)


def _write_metric_sheet(
    worksheet,
    template: TemplateConfig,
    columns: list[ColumnMeta],
    rows: list[RowData],
    values: dict[MetricKey, MetricValue],
    threshold_header_values: dict[str, Any],
    golden_header_values: dict[str, Any],
    threshold_label: str,
    value_number_format: str | None = None,
    threshold_number_format: str | None = None,
) -> None:
    dimension_labels = [item.display_name or item.name for item in template.row_dimensions]
    label_column = len(dimension_labels)
    data_start_column = label_column + 1

    worksheet.cell(1, label_column, "ChainName")
    worksheet.cell(2, label_column, "TestItem")
    worksheet.cell(3, label_column, "RepeatIndex")
    worksheet.cell(4, label_column, threshold_label)
    worksheet.cell(5, label_column, "Golden")

    for offset, column in enumerate(columns):
        cell_column = data_start_column + offset
        worksheet.cell(1, cell_column, column.chain_name)
        worksheet.cell(2, cell_column, column.item_label)
        worksheet.cell(3, cell_column, column.repeat_index)
        threshold_cell = worksheet.cell(4, cell_column, threshold_header_values.get(column.raw_column))
        if threshold_number_format and isinstance(threshold_cell.value, (int, float)):
            threshold_cell.number_format = threshold_number_format
        worksheet.cell(5, cell_column, golden_header_values.get(column.raw_column))
        worksheet.cell(6, cell_column, column.unit)

    for index, label in enumerate(dimension_labels, start=1):
        worksheet.cell(7, index, label)

    data_start_row = 8
    for row_offset, row in enumerate(rows):
        target_row = data_start_row + row_offset
        for index, dimension in enumerate(template.row_dimensions, start=1):
            worksheet.cell(
                target_row,
                index,
                _display_dimension_value(row.dimensions.get(dimension.name)),
            )
        for column_offset, column in enumerate(columns):
            entry = values.get(_row_metric_key(row, column.raw_column))
            cell = worksheet.cell(
                target_row,
                data_start_column + column_offset,
                entry.value if entry else None,
            )
            if entry and value_number_format and isinstance(cell.value, (int, float)):
                cell.number_format = value_number_format
            if entry and entry.is_fail:
                cell.font = FAIL_FONT

    _style_sheet(
        worksheet,
        header_rows=list(range(1, 7)),
        dimension_header_row=7,
        start_column=label_column,
        end_column=data_start_column + len(columns) - 1,
        dimension_count=len(dimension_labels),
    )
    _merge_chain_headers(worksheet, 1, data_start_column, columns)
    _merge_dimension_values(worksheet, data_start_row, len(rows), template)
    _autosize_columns(worksheet)


def _write_point_golden_sheet(
    worksheet,
    template: TemplateConfig,
    columns: list[ColumnMeta],
    rows: list[RowData],
    golden_values: dict[MetricKey, MetricValue],
    golden_header_values: dict[str, Any],
) -> None:
    dimension_labels = [item.display_name or item.name for item in template.row_dimensions]
    label_column = len(dimension_labels)
    data_start_column = label_column + 1
    first_row = 1

    worksheet.cell(first_row, 1, "Matched built golden for each data point")
    worksheet.cell(first_row + 1, label_column, "ChainName")
    worksheet.cell(first_row + 2, label_column, "TestItem")
    worksheet.cell(first_row + 3, label_column, "RepeatIndex")
    worksheet.cell(first_row + 4, label_column, "Golden")

    for offset, column in enumerate(columns):
        cell_column = data_start_column + offset
        worksheet.cell(first_row + 1, cell_column, column.chain_name)
        worksheet.cell(first_row + 2, cell_column, column.item_label)
        worksheet.cell(first_row + 3, cell_column, column.repeat_index)
        worksheet.cell(first_row + 4, cell_column, golden_header_values.get(column.raw_column))
        worksheet.cell(first_row + 5, cell_column, column.unit)

    for index, label in enumerate(dimension_labels, start=1):
        worksheet.cell(first_row + 6, index, label)

    data_start_row = first_row + 7
    for row_offset, row in enumerate(rows):
        target_row = data_start_row + row_offset
        for index, dimension in enumerate(template.row_dimensions, start=1):
            worksheet.cell(
                target_row,
                index,
                _display_dimension_value(row.dimensions.get(dimension.name)),
            )
        for column_offset, column in enumerate(columns):
            entry = golden_values.get(_row_metric_key(row, column.raw_column))
            worksheet.cell(
                target_row,
                data_start_column + column_offset,
                entry.golden_value if entry else None,
            )

    _style_sheet(
        worksheet,
        header_rows=[first_row] + list(range(first_row + 1, first_row + 6)),
        dimension_header_row=first_row + 6,
        start_column=label_column,
        end_column=data_start_column + len(columns) - 1,
        dimension_count=len(dimension_labels),
    )
    if data_start_column + len(columns) - 1 >= 1:
        worksheet.merge_cells(
            start_row=first_row,
            start_column=1,
            end_row=first_row,
            end_column=data_start_column + len(columns) - 1,
        )
    _merge_chain_headers(worksheet, first_row + 1, data_start_column, columns)
    _merge_dimension_values(worksheet, data_start_row, len(rows), template)
    _autosize_columns(worksheet)


def _write_outlier_sheet(
    worksheet,
    template: TemplateConfig,
    columns: list[ColumnMeta],
    rows: list[RowData],
    zscore_values: dict[MetricKey, MetricValue],
    golden_values: dict[MetricKey, MetricValue],
    golden_header_values: dict[str, Any],
    sample_node_orders: dict[str, list[str]],
    summary_rows: list[dict[str, str]],
    outlier_fail_method_override: str | None = None,
) -> None:
    dimension_labels = [item.display_name or item.name for item in template.row_dimensions]
    sample_dimension = template.row_dimensions[0].name if template.row_dimensions else "sample_id"
    fail_mode = outlier_fail_method_override or template.report.outlier_fail_method
    chain_columns = defaultdict(list)
    for column in columns:
        chain_columns[column.chain_name].append(column)

    current_row = 1
    worksheet.cell(1, 1, f"Current Fail Standard: {_format_fail_mode(fail_mode)}")

    rows_by_sample = defaultdict(list)
    for row in rows:
        rows_by_sample[row.dimensions.get(sample_dimension, "")].append(row)

    first_block = True
    for sample_id in sorted(rows_by_sample, key=_natural_sort_key):
        sample_rows = rows_by_sample[sample_id]
        selected_chains = [
            chain_name
            for chain_name in sorted(chain_columns, key=_natural_sort_key)
            if _sample_chain_is_outlier(
                sample_rows,
                chain_columns[chain_name],
                fail_mode,
                template.report.outlier_chain_fail_rule,
                zscore_values,
                golden_values,
            )
        ]
        if not selected_chains:
            continue
        if not first_block:
            current_row += 3
        first_block = False

        label_column = len(dimension_labels) + 1
        data_start_column = label_column + 1
        block_columns = [
            column
            for chain_name in selected_chains
            for column in chain_columns[chain_name]
        ]

        worksheet.cell(current_row, label_column, "ChainName")
        worksheet.cell(current_row + 1, label_column, "TestItem")
        worksheet.cell(current_row + 2, label_column, "RepeatIndex")
        worksheet.cell(current_row + 3, label_column, "Golden")

        for offset, column in enumerate(block_columns):
            cell_column = data_start_column + offset
            worksheet.cell(current_row, cell_column, column.chain_name)
            worksheet.cell(current_row + 1, cell_column, column.item_label)
            worksheet.cell(current_row + 2, cell_column, column.repeat_index)
            worksheet.cell(current_row + 3, cell_column, golden_header_values.get(column.raw_column))
            worksheet.cell(current_row + 4, cell_column, column.unit)

        for index, label in enumerate(dimension_labels, start=1):
            worksheet.cell(current_row + 5, index, label)

        block_data_start = current_row + 6
        sample_rows = sorted(
            sample_rows,
            key=lambda item: _row_sort_key(item, template, sample_node_orders),
        )
        for sample_row_index, row in enumerate(sample_rows):
            row_base = block_data_start + sample_row_index * 4
            for index, dimension in enumerate(template.row_dimensions, start=1):
                worksheet.cell(
                    row_base,
                    index,
                    _display_dimension_value(row.dimensions.get(dimension.name)),
                )

            labels = ["Value", "Zscore", "GoldenDeviation", "Pass/Fail"]
            for label_offset, label in enumerate(labels):
                worksheet.cell(row_base + label_offset, label_column, label)

            failing_chains_for_row: set[str] = set()
            for chain_name in selected_chains:
                chain_fail = _row_chain_is_outlier(
                    row,
                    chain_columns[chain_name],
                    fail_mode,
                    template.report.outlier_chain_fail_rule,
                    zscore_values,
                    golden_values,
                )
                if chain_fail:
                    failing_chains_for_row.add(chain_name)

            for column_offset, column in enumerate(block_columns):
                col_idx = data_start_column + column_offset
                z_entry = zscore_values.get(_row_metric_key(row, column.raw_column))
                golden_entry = golden_values.get(_row_metric_key(row, column.raw_column))
                worksheet.cell(row_base, col_idx, row.values.get(column.raw_column))
                z_cell = worksheet.cell(row_base + 1, col_idx, z_entry.value if z_entry else None)
                g_cell = worksheet.cell(
                    row_base + 2,
                    col_idx,
                    golden_entry.value if golden_entry else None,
                )
                if z_entry and isinstance(z_cell.value, (int, float)):
                    z_cell.number_format = DECIMAL_FORMAT_3
                if golden_entry and isinstance(g_cell.value, (int, float)):
                    g_cell.number_format = PERCENT_FORMAT_2
                pass_fail = _resolve_pass_fail(
                    fail_mode,
                    z_entry,
                    golden_entry,
                )
                pf_cell = worksheet.cell(row_base + 3, col_idx, pass_fail)
                if z_entry and z_entry.is_fail:
                    z_cell.font = FAIL_FONT
                if golden_entry and golden_entry.is_fail:
                    g_cell.font = FAIL_FONT
                if pass_fail == "Fail":
                    pf_cell.font = FAIL_FONT

        _style_sheet(
            worksheet,
            header_rows=list(range(current_row, current_row + 5)),
            dimension_header_row=current_row + 5,
            start_column=label_column,
            end_column=data_start_column + len(block_columns) - 1,
            dimension_count=len(dimension_labels),
        )
        _merge_chain_headers(worksheet, current_row, data_start_column, block_columns)
        _merge_outlier_dimension_values(
            worksheet,
            block_data_start,
            len(sample_rows),
            len(dimension_labels),
        )
        current_row = block_data_start + len(sample_rows) * 4 - 1

    _autosize_columns(worksheet)
    return None


def _write_summary_sheet(
    worksheet,
    template: TemplateConfig,
    summary_rows: list[dict[str, str]],
    ratio_rows: list[dict[str, Any]],
    sample_node_orders: dict[str, list[str]],
) -> None:
    configured_node_orders = _configured_node_orders(template)
    if configured_node_orders:
        worksheet.cell(1, 1, "Configured Node Sequences:")
        for row_offset, node_order in enumerate(configured_node_orders, start=1):
            worksheet.cell(1 + row_offset, 1, f"Sequence {row_offset}")
            for column_offset, node in enumerate(node_order, start=2):
                worksheet.cell(1 + row_offset, column_offset, node)
        start_row = len(configured_node_orders) + 3
    else:
        start_row = 4

    headers = [
        "SampleID",
        "Outlier Node",
        "Outlier Chain",
        "New Outlier? (compared to previous node)",
    ]
    for index, title in enumerate(headers, start=1):
        worksheet.cell(start_row, index, title)

    for row_offset, item in enumerate(
        _sort_summary_rows(summary_rows, sample_node_orders),
        start=1,
    ):
        worksheet.cell(start_row + row_offset, 1, item["sample_id"])
        worksheet.cell(start_row + row_offset, 2, item["node"])
        worksheet.cell(start_row + row_offset, 3, item["chain"])
        worksheet.cell(start_row + row_offset, 4, item["status"])

    ratio_start_row = start_row + len(summary_rows) + 3
    if ratio_rows:
        ratio_headers = [
            "Stat ID",
            "Display Name",
            "Group Dimension",
            "Group Value",
            "Numerator Type",
            "Outlier Sample Count",
            "Total Sample Count",
            "Outlier Ratio",
            "Filter Summary",
        ]
        for index, title in enumerate(ratio_headers, start=1):
            worksheet.cell(ratio_start_row, index, title)
        for row_offset, item in enumerate(ratio_rows, start=1):
            worksheet.cell(ratio_start_row + row_offset, 1, item["id"])
            worksheet.cell(ratio_start_row + row_offset, 2, item["display_name"])
            worksheet.cell(ratio_start_row + row_offset, 3, item["group_by_dimension"])
            worksheet.cell(ratio_start_row + row_offset, 4, item["group_value"])
            worksheet.cell(ratio_start_row + row_offset, 5, item["numerator"])
            worksheet.cell(ratio_start_row + row_offset, 6, item["outlier_sample_count"])
            worksheet.cell(ratio_start_row + row_offset, 7, item["total_sample_count"])
            ratio_cell = worksheet.cell(ratio_start_row + row_offset, 8, item["outlier_ratio"])
            if isinstance(ratio_cell.value, (int, float)):
                ratio_cell.number_format = PERCENT_FORMAT_2
            worksheet.cell(ratio_start_row + row_offset, 9, item["filter_summary"])

    _style_sheet(
        worksheet,
        header_rows=([1] if configured_node_orders else []) + [start_row] + ([ratio_start_row] if ratio_rows else []),
        dimension_header_row=None,
        start_column=1,
        end_column=9 if ratio_rows else 4,
        dimension_count=0,
    )
    _autosize_columns(worksheet)


def _collect_outlier_summary_artifacts(
    template: TemplateConfig,
    columns: list[ColumnMeta],
    rows: list[RowData],
    zscore_values: dict[MetricKey, MetricValue],
    golden_values: dict[MetricKey, MetricValue],
    sample_node_orders: dict[str, list[str]],
    outlier_fail_method_override: str | None = None,
) -> dict[str, list[dict[str, Any]]]:
    sample_dimension = template.row_dimensions[0].name if template.row_dimensions else "sample_id"
    fail_mode = outlier_fail_method_override or template.report.outlier_fail_method
    chain_columns = defaultdict(list)
    for column in columns:
        chain_columns[column.chain_name].append(column)

    rows_by_sample = defaultdict(list)
    for row in rows:
        rows_by_sample[row.dimensions.get(sample_dimension, "")].append(row)

    summary_events: list[dict[str, Any]] = []
    for sample_id in sorted(rows_by_sample, key=_natural_sort_key):
        sample_rows = sorted(
            rows_by_sample[sample_id],
            key=lambda item: _row_sort_key(item, template, sample_node_orders),
        )
        selected_chains = [
            chain_name
            for chain_name in sorted(chain_columns, key=_natural_sort_key)
            if _sample_chain_is_outlier(
                sample_rows,
                chain_columns[chain_name],
                fail_mode,
                template.report.outlier_chain_fail_rule,
                zscore_values,
                golden_values,
            )
        ]
        if not selected_chains:
            continue

        chain_fail_details = _build_sample_chain_fail_details(
            sample_rows,
            selected_chains,
            chain_columns,
            fail_mode,
            template.report.outlier_chain_fail_rule,
            zscore_values,
            golden_values,
        )
        node_predecessors = _build_sample_node_predecessors(
            {row.dimensions.get("reliability_node", "") for row in sample_rows if row.dimensions.get("reliability_node", "")},
            _configured_node_orders(template),
        )

        for chain_name in sorted(selected_chains, key=_natural_sort_key):
            failed_nodes = sorted(
                chain_fail_details.get(chain_name, {}).keys(),
                key=lambda node: _node_sort_key(node, sample_node_orders.get(sample_id, [])),
            )
            for node_value in failed_nodes:
                predecessor_nodes = node_predecessors.get(node_value, set())
                status = (
                    "Existed"
                    if any(
                        predecessor in chain_fail_details.get(chain_name, {})
                        for predecessor in predecessor_nodes
                    )
                    else "New"
                )
                detail = chain_fail_details.get(chain_name, {}).get(node_value, {})
                summary_events.append(
                    {
                        "sample_id": sample_id,
                        "node": node_value,
                        "chain": chain_name,
                        "status": status,
                        "dimensions": dict(
                            next(
                                (
                                    row.dimensions
                                    for row in sample_rows
                                    if row.dimensions.get("reliability_node", "") == node_value
                                ),
                                {},
                            )
                        ),
                        "raw_columns": sorted(detail.get("raw_columns", set())),
                        "logical_metrics": sorted(detail.get("logical_metrics", set())),
                    }
                )

    summary_rows = _dedupe_summary_rows(
        [
            {
                "sample_id": item["sample_id"],
                "node": item["node"],
                "chain": item["chain"],
                "status": item["status"],
            }
            for item in summary_events
        ],
        template,
        sample_node_orders,
    )
    ratio_rows = _build_outlier_ratio_rows(
        template.report.outlier_ratio_stats,
        summary_events,
        rows,
        sample_node_orders,
    )
    return {
        "summary_rows": summary_rows,
        "ratio_rows": ratio_rows,
    }


def _row_sort_key(
    row: RowData,
    template: TemplateConfig,
    sample_node_orders: dict[str, list[str]] | None = None,
) -> tuple[Any, ...]:
    sample_id = row.dimensions.get("sample_id", "")
    node_value = row.dimensions.get("reliability_node", "")
    repeat_value = row.dimensions.get("repeat_id", "")
    node_order = []
    if sample_node_orders is not None:
        node_order = sample_node_orders.get(sample_id, [])
    if not node_order:
        configured = _configured_node_orders(template)
        node_order = configured[0] if configured else []
    return (
        _natural_sort_key(sample_id),
        _node_sort_key(node_value, node_order),
        _safe_int(repeat_value),
        tuple(_natural_sort_key(row.dimensions.get(item.name, "")) for item in template.row_dimensions[3:]),
    )


def _parse_measurement_column(raw_column: str) -> dict[str, Any]:
    parts = [item for item in raw_column.split("_") if item]
    metric = parts[0] if parts else raw_column
    chain_name = next((item for item in parts if item.startswith("Link")), raw_column)
    repeat_index = 1
    condition_parts: list[str] = []
    if parts and parts[-1].isdigit():
        repeat_index = int(parts[-1])
        condition_parts = parts[2:-1]
    else:
        condition_parts = parts[2:]
    if condition_parts and len(condition_parts) == 1 and condition_parts[0].isdigit():
        condition_parts = []
        repeat_index = int(parts[-1])
    item_label = metric
    if condition_parts:
        item_label = f"{metric}_{'_'.join(condition_parts)}"
    return {
        "chain_name": chain_name,
        "item_label": item_label,
        "repeat_index": repeat_index,
        "metric_type": metric,
        "condition": "_".join(condition_parts),
    }


def _resolve_threshold(config: ThresholdConfig, candidates: list[str]) -> float:
    for candidate in candidates:
        if candidate in config.overrides:
            return config.overrides[candidate]
    return config.default


def _resolve_golden_value(template: TemplateConfig, candidates: list[str]) -> float | None:
    for candidate in candidates:
        if candidate in template.report.golden_values:
            return template.report.golden_values[candidate]
    return None


def _resolve_built_relative_threshold(entry: AnomalyResult) -> float | None:
    if entry.center is None or entry.lower_bound is None or entry.upper_bound is None:
        return None
    if entry.center == 0:
        return None
    return max(
        abs(entry.upper_bound - entry.center),
        abs(entry.center - entry.lower_bound),
    ) / abs(entry.center)


def _resolve_built_absolute_threshold(entry: AnomalyResult) -> float | None:
    if entry.center is None or entry.lower_bound is None or entry.upper_bound is None:
        return None
    return max(
        abs(entry.upper_bound - entry.center),
        abs(entry.center - entry.lower_bound),
    )


def _sample_chain_is_outlier(
    sample_rows: list[RowData],
    columns: list[ColumnMeta],
    fail_mode: str,
    fail_rule: str,
    zscore_values: dict[MetricKey, MetricValue],
    golden_values: dict[MetricKey, MetricValue],
) -> bool:
    return any(
        _row_chain_is_outlier(row, columns, fail_mode, fail_rule, zscore_values, golden_values)
        for row in sample_rows
    )


def _row_chain_is_outlier(
    row: RowData,
    columns: list[ColumnMeta],
    fail_mode: str,
    fail_rule: str,
    zscore_values: dict[MetricKey, MetricValue],
    golden_values: dict[MetricKey, MetricValue],
) -> bool:
    statuses: list[bool] = []
    for column in columns:
        z_entry = zscore_values.get(_row_metric_key(row, column.raw_column))
        golden_entry = golden_values.get(_row_metric_key(row, column.raw_column))
        statuses.append(_is_metric_fail(fail_mode, z_entry, golden_entry))
    if not statuses:
        return False
    if fail_rule == "all_fail":
        return all(statuses)
    return any(statuses)


def _resolve_pass_fail(
    fail_mode: str,
    z_entry: MetricValue | None,
    golden_entry: MetricValue | None,
) -> str:
    return "Fail" if _is_metric_fail(fail_mode, z_entry, golden_entry) else "Pass"


def _is_metric_fail(
    fail_mode: str,
    z_entry: MetricValue | None,
    golden_entry: MetricValue | None,
) -> bool:
    z_fail = bool(z_entry and z_entry.is_fail)
    golden_fail = bool(golden_entry and golden_entry.is_fail)
    if fail_mode == "modified_z_score":
        return z_fail
    if fail_mode == "golden_deviation":
        return golden_fail
    if fail_mode == "zscore_and_golden":
        return z_fail and golden_fail
    if fail_mode == "zscore_or_golden":
        return z_fail or golden_fail
    raise ValueError(f"Unsupported fail mode: {fail_mode}")


def _style_sheet(
    worksheet,
    header_rows: list[int],
    dimension_header_row: int | None,
    start_column: int,
    end_column: int,
    dimension_count: int,
) -> None:
    for row_index in header_rows:
        for col_index in range(start_column, end_column + 1):
            cell = worksheet.cell(row_index, col_index)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
    if dimension_header_row is not None:
        for col_index in range(1, max(end_column, dimension_count) + 1):
            cell = worksheet.cell(dimension_header_row, col_index)
            cell.font = HEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if cell.alignment == Alignment():
                cell.alignment = CENTER
            if cell.border == Border():
                cell.border = THIN_BORDER


def _merge_chain_headers(worksheet, header_row: int, data_start_column: int, columns: list[ColumnMeta]) -> None:
    if not columns:
        return
    start = data_start_column
    current = columns[0].chain_name
    for offset, column in enumerate(columns[1:], start=1):
        absolute_column = data_start_column + offset
        if column.chain_name != current:
            if absolute_column - 1 > start:
                worksheet.merge_cells(
                    start_row=header_row,
                    start_column=start,
                    end_row=header_row,
                    end_column=absolute_column - 1,
                )
            start = absolute_column
            current = column.chain_name
    end_column = data_start_column + len(columns) - 1
    if end_column > start:
        worksheet.merge_cells(
            start_row=header_row,
            start_column=start,
            end_row=header_row,
            end_column=end_column,
        )


def _merge_dimension_values(
    worksheet,
    start_row: int,
    row_count: int,
    template: TemplateConfig,
) -> None:
    if row_count <= 1:
        return
    dimension_names = [item.name for item in template.row_dimensions]
    for col_index, _dimension in enumerate(dimension_names, start=1):
        group_start = start_row
        previous = worksheet.cell(start_row, col_index).value
        for row_index in range(start_row + 1, start_row + row_count + 1):
            current = worksheet.cell(row_index, col_index).value if row_index < start_row + row_count else None
            should_merge = (
                row_index < start_row + row_count
                and current == previous
                and _same_prefix_values(worksheet, group_start, row_index, col_index)
            )
            if should_merge:
                continue
            if row_index - group_start > 1 and previous not in (None, ""):
                worksheet.merge_cells(
                    start_row=group_start,
                    start_column=col_index,
                    end_row=row_index - 1,
                    end_column=col_index,
                )
            group_start = row_index
            previous = current


def _merge_outlier_dimension_values(
    worksheet,
    start_row: int,
    row_count: int,
    dimension_count: int,
) -> None:
    if row_count <= 0:
        return

    block_height = 4
    end_row_exclusive = start_row + row_count * block_height

    for col_index in range(1, dimension_count + 1):
        group_start = start_row
        previous = worksheet.cell(start_row, col_index).value

        for row_index in range(start_row + block_height, end_row_exclusive + block_height, block_height):
            current = (
                worksheet.cell(row_index, col_index).value
                if row_index < end_row_exclusive
                else None
            )
            should_merge = (
                row_index < end_row_exclusive
                and current == previous
                and _same_prefix_values(worksheet, group_start, row_index, col_index)
            )
            if should_merge:
                continue

            if previous not in (None, ""):
                worksheet.merge_cells(
                    start_row=group_start,
                    start_column=col_index,
                    end_row=row_index - 1,
                    end_column=col_index,
                )

            group_start = row_index
            previous = current


def _same_prefix_values(worksheet, start_row: int, current_row: int, current_column: int) -> bool:
    for col_index in range(1, current_column):
        if worksheet.cell(start_row, col_index).value != worksheet.cell(current_row, col_index).value:
            return False
    return True


def _find_any_metric_value(raw_column: str, values: dict[MetricKey, MetricValue]) -> MetricValue | None:
    for (_dataset_id, _row_number, column), value in values.items():
        if column == raw_column:
            return value
    return None


def _find_any_golden_value(raw_column: str, values: dict[MetricKey, MetricValue]) -> float | None:
    entry = _find_any_metric_value(raw_column, values)
    return entry.golden_value if entry else None


def _build_column_display_values(
    values: dict[MetricKey, MetricValue],
    attribute: str,
) -> dict[str, Any]:
    grouped: dict[str, list[Any]] = defaultdict(list)
    for (_dataset_id, _row_number, raw_column), metric_value in values.items():
        value = getattr(metric_value, attribute)
        if value is None:
            continue
        grouped[raw_column].append(value)

    display_values: dict[str, Any] = {}
    for raw_column, items in grouped.items():
        normalized = _collapse_display_values(items)
        display_values[raw_column] = normalized
    return display_values


def _collapse_display_values(items: list[Any]) -> Any:
    if not items:
        return None
    if all(isinstance(item, (int, float)) for item in items):
        rounded = {round(float(item), 12) for item in items}
        if len(rounded) == 1:
            return items[0]
        return "Varies"
    unique = {str(item) for item in items}
    if len(unique) == 1:
        return items[0]
    return "Varies"


def _has_variable_metric_attribute(
    values: dict[MetricKey, MetricValue],
    attribute: str,
) -> bool:
    grouped: dict[str, set[Any]] = defaultdict(set)
    for (_dataset_id, _row_number, raw_column), metric_value in values.items():
        value = getattr(metric_value, attribute)
        if value is None:
            continue
        if isinstance(value, float):
            grouped[raw_column].add(round(value, 12))
        else:
            grouped[raw_column].add(value)
    return any(len(items) > 1 for items in grouped.values())


def _autosize_columns(worksheet) -> None:
    for index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 10), 24)


def _dedupe_summary_rows(
    rows: list[dict[str, str]],
    template: TemplateConfig,
    sample_node_orders: dict[str, list[str]],
) -> list[dict[str, str]]:
    seen: set[tuple[str, str, str]] = set()
    deduped: list[dict[str, str]] = []
    for item in _sort_summary_rows(rows, sample_node_orders):
        key = (item["sample_id"], item["node"], item["chain"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def _build_sample_chain_fail_details(
    sample_rows: list[RowData],
    selected_chains: list[str],
    chain_columns: dict[str, list[ColumnMeta]],
    fail_mode: str,
    fail_rule: str,
    zscore_values: dict[MetricKey, MetricValue],
    golden_values: dict[MetricKey, MetricValue],
) -> dict[str, dict[str, dict[str, set[str]]]]:
    fail_details: dict[str, dict[str, dict[str, set[str]]]] = defaultdict(dict)
    for row in sample_rows:
        node_value = row.dimensions.get("reliability_node", "")
        if not node_value:
            continue
        for chain_name in selected_chains:
            failing_columns: list[ColumnMeta] = []
            statuses: list[bool] = []
            for column in chain_columns[chain_name]:
                z_entry = zscore_values.get(_row_metric_key(row, column.raw_column))
                golden_entry = golden_values.get(_row_metric_key(row, column.raw_column))
                is_fail = _is_metric_fail(fail_mode, z_entry, golden_entry)
                statuses.append(is_fail)
                if is_fail:
                    failing_columns.append(column)
            if not statuses:
                continue
            chain_fail = all(statuses) if fail_rule == "all_fail" else any(statuses)
            if not chain_fail:
                continue
            fail_details[chain_name][node_value] = {
                "raw_columns": {column.raw_column for column in failing_columns},
                "logical_metrics": {column.logical_metric for column in failing_columns},
            }
    return fail_details


def _build_sample_node_predecessors(
    observed_nodes: set[str],
    configured_node_orders: list[list[str]],
) -> dict[str, set[str]]:
    predecessors: dict[str, set[str]] = defaultdict(set)
    for node_order in configured_node_orders:
        for previous_node, current_node in zip(node_order, node_order[1:]):
            if previous_node in observed_nodes and current_node in observed_nodes:
                predecessors[current_node].add(previous_node)
    return predecessors


def _build_outlier_ratio_rows(
    configs: list[OutlierRatioStatConfig],
    summary_events: list[dict[str, Any]],
    rows: list[RowData],
    sample_node_orders: dict[str, list[str]],
) -> list[dict[str, Any]]:
    if not configs:
        return []
    merged_node_order = _merge_node_orders(sample_node_orders.values())
    rows_by_dimension: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for row in rows:
        sample_id = row.dimensions.get("sample_id", "")
        if not sample_id:
            continue
        for config in configs:
            group_value = str(row.dimensions.get(config.group_by_dimension, "") or "")
            if group_value:
                rows_by_dimension[config.id][group_value].add(sample_id)

    ratio_rows: list[dict[str, Any]] = []
    for config in configs:
        sample_ids_by_group = rows_by_dimension.get(config.id, {})
        for group_value in sorted(
            sample_ids_by_group,
            key=(
                (lambda item: _node_sort_key(item, merged_node_order))
                if config.group_by_dimension == "reliability_node"
                else _natural_sort_key
            ),
        ):
            total_sample_count = len(sample_ids_by_group[group_value])
            matched_sample_ids = {
                str(event["sample_id"])
                for event in summary_events
                if str(event.get("dimensions", {}).get(config.group_by_dimension, "") or "") == group_value
                and _summary_event_matches_ratio_config(event, config)
                and (
                    config.numerator != "new_outlier_samples"
                    or str(event.get("status", "")) == "New"
                )
            }
            ratio_rows.append(
                {
                    "id": config.id,
                    "display_name": config.display_name or config.id,
                    "group_by_dimension": config.group_by_dimension,
                    "group_value": group_value,
                    "numerator": config.numerator,
                    "outlier_sample_count": len(matched_sample_ids),
                    "total_sample_count": total_sample_count,
                    "outlier_ratio": (
                        len(matched_sample_ids) / total_sample_count
                        if total_sample_count
                        else 0.0
                    ),
                    "filter_summary": _format_ratio_filter_summary(config),
                }
            )
    return sorted(
        ratio_rows,
        key=lambda item: (
            _natural_sort_key(str(item["display_name"])),
            _node_sort_key(
                str(item["group_value"]),
                merged_node_order,
            )
            if str(item["group_by_dimension"]) == "reliability_node"
            else _natural_sort_key(str(item["group_value"])),
        ),
    )


def _summary_event_matches_ratio_config(
    event: dict[str, Any],
    config: OutlierRatioStatConfig,
) -> bool:
    if config.chains and str(event.get("chain", "")) not in config.chains:
        return False
    if config.raw_columns:
        event_raw_columns = set(event.get("raw_columns", []))
        if not event_raw_columns.intersection(config.raw_columns):
            return False
    if config.logical_metrics:
        event_logical_metrics = set(event.get("logical_metrics", []))
        if not event_logical_metrics.intersection(config.logical_metrics):
            return False
    return True


def _format_ratio_filter_summary(config: OutlierRatioStatConfig) -> str:
    items: list[str] = []
    if config.chains:
        items.append("chain=" + ",".join(config.chains))
    if config.raw_columns:
        items.append("raw_column=" + ",".join(config.raw_columns))
    if config.logical_metrics:
        items.append("logical_metric=" + ",".join(config.logical_metrics))
    return " | ".join(items) if items else "No filters"


def _merge_node_orders(node_orders: Any) -> list[str]:
    merged: list[str] = []
    for node_order in node_orders:
        for node in node_order:
            if node not in merged:
                merged.append(node)
    return merged


def _format_fail_mode(value: str) -> str:
    if value == "modified_z_score":
        return "Zscore"
    if value == "golden_deviation":
        return "GoldenDeviation"
    if value == "zscore_and_golden":
        return "Zscore AND GoldenDeviation"
    if value == "zscore_or_golden":
        return "Zscore OR GoldenDeviation"
    return value


def _build_sample_node_orders(
    rows: list[RowData],
    template: TemplateConfig,
) -> dict[str, list[str]]:
    configured = _configured_node_orders(template)
    if not configured:
        return {}
    nodes_by_sample: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        sample_id = row.dimensions.get("sample_id", "")
        node_value = row.dimensions.get("reliability_node", "")
        if sample_id and node_value:
            nodes_by_sample[sample_id].add(node_value)
    return {
        sample_id: _resolve_sample_node_order(observed_nodes, configured)
        for sample_id, observed_nodes in nodes_by_sample.items()
    }


def _configured_node_orders(template: TemplateConfig) -> list[list[str]]:
    if template.report.node_orders:
        return [list(item) for item in template.report.node_orders if item]
    if template.report.node_order:
        return [list(template.report.node_order)]
    return []


def _resolve_sample_node_order(
    observed_nodes: set[str],
    configured_node_orders: list[list[str]],
) -> list[str]:
    best_order = configured_node_orders[0]
    best_score = _node_order_match_score(observed_nodes, best_order)
    for candidate in configured_node_orders[1:]:
        score = _node_order_match_score(observed_nodes, candidate)
        if score < best_score:
            best_order = candidate
            best_score = score
    return list(best_order)


def _node_order_match_score(
    observed_nodes: set[str],
    node_order: list[str],
) -> tuple[int, int, tuple[Any, ...]]:
    configured_nodes = set(node_order)
    unknown_count = len([node for node in observed_nodes if node not in configured_nodes])
    unused_configured_count = len([node for node in node_order if node not in observed_nodes])
    return (
        unknown_count,
        unused_configured_count,
        tuple(_natural_sort_key(node) for node in node_order),
    )


def _node_sort_key(value: str, node_order: list[str]) -> tuple[int, Any]:
    if value in node_order:
        return (0, node_order.index(value))
    return (1, _natural_sort_key(value))


def _sort_summary_rows(
    rows: list[dict[str, str]],
    sample_node_orders: dict[str, list[str]],
) -> list[dict[str, str]]:
    return sorted(
        rows,
        key=lambda payload: (
            _natural_sort_key(payload["sample_id"]),
            _node_sort_key(
                payload["node"],
                sample_node_orders.get(payload["sample_id"], []),
            ),
            _natural_sort_key(payload["chain"]),
        ),
    )


def _natural_sort_key(value: str) -> tuple[Any, ...]:
    parts = re.split(r"(\d+)", value or "")
    key: list[Any] = []
    for part in parts:
        if not part:
            continue
        if part.isdigit():
            key.append(int(part))
        else:
            key.append(part.lower())
    return tuple(key)


def _safe_int(value: str | None) -> int:
    if value is None:
        return 0
    match = re.search(r"\d+", str(value))
    if match:
        return int(match.group())
    return 0


def _display_dimension_value(value: str | None) -> Any:
    if value is None:
        return None
    text = str(value)
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return float(text)
    return text


def _metric_sort_key(metric_type: str) -> tuple[int, str]:
    order = {"R": 0, "I": 1, "V": 2, "C": 3}
    return (order.get(metric_type, 99), metric_type)


def _condition_sort_key(condition: str) -> tuple[int, float, str]:
    if not condition:
        return (1, 0.0, "")
    match = re.fullmatch(r"(\d+(?:p\d+)?)([A-Za-z]+)", condition)
    if not match:
        return (0, 0.0, condition)
    numeric = float(match.group(1).replace("p", "."))
    unit = match.group(2).lower()
    return (0, numeric, unit)
