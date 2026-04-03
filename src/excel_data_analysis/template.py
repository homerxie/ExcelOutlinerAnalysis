from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from .models import (
    AnalysisGroup,
    GoldenReferenceDefaults,
    OutlierRatioStatConfig,
    ReportConfig,
    RowDimension,
    RowDimensionSource,
    TemplateConfig,
    ThresholdConfig,
)
from .output_paths import resolve_output_path

TEMPLATE_WORKBOOK_SHEETS = {
    "info": "template_info",
    "row_dimensions": "row_dimensions",
    "analysis_groups": "analysis_groups",
    "node_orders": "node_orders",
    "golden_values": "golden_values",
    "zscore_thresholds": "zscore_thresholds",
    "golden_deviation_thresholds": "golden_deviation_thresholds",
    "outlier_ratio_stats": "outlier_ratio_stats",
    "readme": "README",
}


def load_template(path: str | Path) -> TemplateConfig:
    template_path = Path(path)
    suffix = template_path.suffix.lower()
    if suffix == ".json":
        payload = json.loads(template_path.read_text(encoding="utf-8"))
        return _parse_template(payload)
    if suffix in {".xlsx", ".xlsm"}:
        payload = _load_template_workbook_payload(template_path)
        return _parse_template(payload)
    raise ValueError(f"Unsupported template format: {template_path.suffix}")


def save_template(
    template: TemplateConfig,
    path: str | Path,
    if_exists: str = "overwrite",
) -> Path:
    template_path = resolve_output_path(path, if_exists=if_exists)
    payload = template_to_payload(template)
    suffix = template_path.suffix.lower()
    if suffix == ".json":
        template_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        return template_path
    if suffix == ".xlsx":
        _save_template_workbook(payload, template_path)
        return template_path
    raise ValueError(f"Unsupported template output format: {template_path.suffix}")


def validate_template_file(path: str | Path) -> TemplateConfig:
    return load_template(path)


def summarize_template(template: TemplateConfig) -> dict[str, Any]:
    return {
        "name": template.name,
        "row_dimension_count": len(template.row_dimensions),
        "analysis_group_count": len(template.analysis_groups),
        "golden_reference_default_filter_count": len(
            template.golden_reference_defaults.filters
        ),
        "golden_value_count": len(template.report.golden_values),
        "zscore_threshold_override_count": len(
            template.report.zscore_thresholds.overrides
        ),
        "golden_deviation_threshold_override_count": len(
            template.report.golden_deviation_thresholds.overrides
        ),
        "outlier_ratio_stat_count": len(template.report.outlier_ratio_stats),
        "node_order_count": len(_effective_node_orders(template.report)[0])
        if _effective_node_orders(template.report)
        else 0,
        "node_sequence_count": len(_effective_node_orders(template.report)),
        "measurement_header_row": template.measurement_header_row,
        "row_header_row": template.row_header_row,
        "unit_row": template.unit_row,
        "data_start_row": template.data_start_row,
    }


def template_to_payload(template: TemplateConfig) -> dict[str, Any]:
    report_payload: dict[str, Any] = {
        "golden_values": {
            key: _serialize_number(value)
            for key, value in template.report.golden_values.items()
        },
        "zscore_thresholds": _threshold_config_to_payload(
            template.report.zscore_thresholds
        ),
        "golden_deviation_thresholds": _threshold_config_to_payload(
            template.report.golden_deviation_thresholds
        ),
        "outlier_fail_method": template.report.outlier_fail_method,
        "outlier_chain_fail_rule": template.report.outlier_chain_fail_rule,
        "outlier_ratio_stats": [
            _outlier_ratio_stat_to_payload(item)
            for item in template.report.outlier_ratio_stats
        ],
    }
    if not template.report.golden_deviation_thresholds.overrides:
        report_payload.pop("golden_deviation_thresholds", None)
    node_orders = _effective_node_orders(template.report)
    if len(node_orders) <= 1:
        report_payload["node_order"] = list(node_orders[0]) if node_orders else []
    else:
        report_payload["node_orders"] = [list(item) for item in node_orders]

    return {
        "name": template.name,
        "header_row": template.header_row,
        "measurement_header_row": template.measurement_header_row,
        "row_header_row": template.row_header_row,
        "unit_row": template.unit_row,
        "data_start_row": template.data_start_row,
        "row_dimensions": [
            _row_dimension_to_payload(item) for item in template.row_dimensions
        ],
        "analysis_groups": [
            _analysis_group_to_payload(item) for item in template.analysis_groups
        ],
        "golden_reference_defaults": _golden_reference_defaults_to_payload(
            template.golden_reference_defaults
        ),
        "report": report_payload,
    }


def _row_dimension_to_payload(item: RowDimension) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "name": item.name,
        "sources": [
            _row_dimension_source_to_payload(source) for source in item.sources
        ],
    }
    if item.optional:
        payload["optional"] = True
    if item.display_name is not None:
        payload["display_name"] = item.display_name
    if item.default_value is not None:
        payload["default_value"] = item.default_value
    return payload


def _row_dimension_source_to_payload(source: RowDimensionSource) -> dict[str, Any]:
    payload: dict[str, Any] = {"column": source.column}
    if source.split_delimiters:
        payload["split_delimiters"] = list(source.split_delimiters)
    if source.split_position is not None:
        payload["split_position"] = source.split_position
    if source.skip_empty_parts:
        payload["skip_empty_parts"] = True
    return payload


def _analysis_group_to_payload(item: AnalysisGroup) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "id": item.id,
        "display_name": item.display_name,
        "columns": list(item.columns),
        "analysis_mode": item.analysis_mode,
    }
    if item.presentation_group is not None:
        payload["presentation_group"] = item.presentation_group
    if item.unit is not None:
        payload["unit"] = item.unit
    return payload


def _threshold_config_to_payload(config: ThresholdConfig) -> dict[str, Any]:
    payload: dict[str, Any] = {"default": _serialize_number(config.default)}
    if config.overrides:
        payload["overrides"] = {
            key: _serialize_number(value) for key, value in config.overrides.items()
        }
    return payload


def _golden_reference_defaults_to_payload(
    defaults: GoldenReferenceDefaults,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "reference_dimensions": list(defaults.reference_dimensions),
        "filters": dict(defaults.filters),
        "center_method": defaults.center_method,
        "threshold_mode": defaults.threshold_mode,
    }
    if defaults.relative_limit is not None:
        payload["relative_limit"] = _serialize_number(defaults.relative_limit)
    if defaults.sigma_multiplier is not None:
        payload["sigma_multiplier"] = _serialize_number(defaults.sigma_multiplier)
    return payload


def _resolve_golden_deviation_default(
    golden_reference_payload: dict[str, Any],
) -> float:
    if golden_reference_payload.get("relative_limit") not in {None, ""}:
        return float(golden_reference_payload["relative_limit"])
    return 0.2


def _outlier_ratio_stat_to_payload(item: OutlierRatioStatConfig) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "id": item.id,
        "group_by_dimension": item.group_by_dimension,
        "numerator": item.numerator,
    }
    if item.display_name is not None:
        payload["display_name"] = item.display_name
    if item.chains:
        payload["chains"] = list(item.chains)
    if item.raw_columns:
        payload["raw_columns"] = list(item.raw_columns)
    if item.logical_metrics:
        payload["logical_metrics"] = list(item.logical_metrics)
    return payload


def _parse_template(payload: dict[str, Any]) -> TemplateConfig:
    row_dimensions = [
        RowDimension(
            name=item["name"],
            sources=[
                RowDimensionSource(
                    column=source["column"],
                    split_delimiters=[
                        str(delimiter)
                        for delimiter in source.get("split_delimiters", [])
                        if str(delimiter)
                    ],
                    split_position=source.get("split_position"),
                    skip_empty_parts=source.get("skip_empty_parts", False),
                )
                for source in item.get("sources", [])
            ],
            optional=item.get("optional", False),
            display_name=item.get("display_name"),
            default_value=item.get("default_value"),
        )
        for item in payload.get("row_dimensions", [])
    ]
    analysis_groups = [
        AnalysisGroup(
            id=item["id"],
            display_name=item.get("display_name", item["id"]),
            columns=item["columns"],
            analysis_mode=item["analysis_mode"],
            presentation_group=item.get("presentation_group"),
            unit=item.get("unit"),
        )
        for item in payload.get("analysis_groups", [])
    ]
    report_payload = payload.get("report", {})
    outlier_ratio_stats = [
        OutlierRatioStatConfig(
            id=item["id"],
            display_name=item.get("display_name"),
            group_by_dimension=item.get("group_by_dimension", "reliability_node"),
            numerator=item.get("numerator", "new_outlier_samples"),
            chains=[str(value).strip() for value in item.get("chains", []) if str(value).strip()],
            raw_columns=[str(value).strip() for value in item.get("raw_columns", []) if str(value).strip()],
            logical_metrics=[str(value).strip() for value in item.get("logical_metrics", []) if str(value).strip()],
        )
        for item in report_payload.get("outlier_ratio_stats", [])
    ]
    node_orders = _parse_node_orders_payload(report_payload)
    golden_reference_payload = payload.get("golden_reference_defaults", {})
    relative_limit_raw = golden_reference_payload.get("relative_limit")
    sigma_multiplier_raw = golden_reference_payload.get("sigma_multiplier")
    golden_default_threshold = _resolve_golden_deviation_default(
        golden_reference_payload,
    )
    template = TemplateConfig(
        name=payload["name"],
        header_row=payload.get("header_row", 0),
        measurement_header_row=payload.get("measurement_header_row"),
        row_header_row=payload.get("row_header_row"),
        unit_row=payload.get("unit_row"),
        data_start_row=payload.get("data_start_row"),
        row_dimensions=row_dimensions,
        analysis_groups=analysis_groups,
        golden_reference_defaults=GoldenReferenceDefaults(
            reference_dimensions=[
                str(item).strip()
                for item in golden_reference_payload.get("reference_dimensions", [])
                if str(item).strip()
            ],
            filters={
                str(key): str(value)
                for key, value in golden_reference_payload.get("filters", {}).items()
            },
            center_method=golden_reference_payload.get("center_method", "mean"),
            threshold_mode=golden_reference_payload.get("threshold_mode", "relative"),
            relative_limit=(
                golden_default_threshold
                if relative_limit_raw in {None, ""}
                else _parse_optional_float_value(relative_limit_raw)
            ),
            sigma_multiplier=(
                3.0
                if sigma_multiplier_raw in {None, ""}
                else _parse_optional_float_value(sigma_multiplier_raw)
            ),
        ),
        report=ReportConfig(
            golden_values={
                key: float(value)
                for key, value in report_payload.get("golden_values", {}).items()
            },
            zscore_thresholds=_parse_threshold_config(
                report_payload.get("zscore_thresholds"),
                default=3.5,
            ),
            golden_deviation_thresholds=_parse_threshold_config(
                report_payload.get("golden_deviation_thresholds"),
                default=golden_default_threshold,
            ),
            node_order=list(node_orders[0]) if node_orders else [],
            node_orders=node_orders,
            outlier_fail_method=report_payload.get(
                "outlier_fail_method", "modified_z_score"
            ),
            outlier_chain_fail_rule=report_payload.get(
                "outlier_chain_fail_rule", "any_fail"
            ),
            outlier_ratio_stats=outlier_ratio_stats,
        ),
    )
    _validate_template(template)
    return template


def _parse_threshold_config(
    payload: dict | float | int | None, default: float
) -> ThresholdConfig:
    if payload is None:
        return ThresholdConfig(default=default)
    if isinstance(payload, (int, float)):
        return ThresholdConfig(default=float(payload))
    return ThresholdConfig(
        default=float(payload.get("default", default)),
        overrides={
            str(key): float(value) for key, value in payload.get("overrides", {}).items()
        },
    )


def _parse_node_orders_payload(report_payload: dict[str, Any]) -> list[list[str]]:
    if "node_orders" in report_payload and report_payload.get("node_orders") is not None:
        payload = report_payload.get("node_orders")
        if not isinstance(payload, list):
            raise ValueError("report.node_orders must be a list of node sequences.")
        node_orders: list[list[str]] = []
        for index, item in enumerate(payload, start=1):
            if not isinstance(item, list):
                raise ValueError(
                    f"report.node_orders[{index}] must be a list of node names."
                )
            nodes = [str(node).strip() for node in item if str(node).strip()]
            node_orders.append(nodes)
        return node_orders
    legacy_payload = report_payload.get("node_order", [])
    if legacy_payload and isinstance(legacy_payload, list) and isinstance(legacy_payload[0], list):
        return [
            [str(node).strip() for node in item if str(node).strip()]
            for item in legacy_payload
        ]
    return [[str(item).strip() for item in legacy_payload if str(item).strip()]]


def _validate_template(template: TemplateConfig) -> None:
    if template.measurement_header_row is None:
        template.measurement_header_row = template.header_row
    if template.row_header_row is None:
        template.row_header_row = template.measurement_header_row

    seen_dimensions: set[str] = set()
    for dimension in template.row_dimensions:
        if dimension.name in seen_dimensions:
            raise ValueError(f"Duplicate row dimension: {dimension.name}")
        seen_dimensions.add(dimension.name)
        if not dimension.sources:
            raise ValueError(f"Row dimension {dimension.name} must define at least one source.")
        for source_index, source in enumerate(dimension.sources, start=1):
            if not source.column:
                raise ValueError(
                    f"Row dimension {dimension.name} source {source_index} must define a column."
                )
            if source.split_position is not None and not source.split_delimiters:
                raise ValueError(
                    f"Row dimension {dimension.name} source {source_index} uses split_position and must define split_delimiters."
                )

    seen_groups: set[str] = set()
    seen_columns: set[str] = set()
    for group in template.analysis_groups:
        if group.id in seen_groups:
            raise ValueError(f"Duplicate analysis group id: {group.id}")
        seen_groups.add(group.id)
        if group.analysis_mode not in {"pooled_columns", "per_column"}:
            raise ValueError(
                f"Unsupported analysis mode for group {group.id}: {group.analysis_mode}"
            )
        for column in group.columns:
            if column in seen_columns:
                raise ValueError(
                    f"Column {column} is mapped by more than one analysis group."
                )
            seen_columns.add(column)

    if template.report.outlier_fail_method not in {
        "modified_z_score",
        "golden_deviation",
        "zscore_and_golden",
        "zscore_or_golden",
    }:
        raise ValueError(
            "report.outlier_fail_method must be one of modified_z_score, golden_deviation, zscore_and_golden, or zscore_or_golden."
        )

    if template.report.outlier_chain_fail_rule not in {"any_fail", "all_fail"}:
        raise ValueError(
            "report.outlier_chain_fail_rule must be one of any_fail or all_fail."
        )

    for stat in template.report.outlier_ratio_stats:
        if stat.numerator not in {"new_outlier_samples", "outlier_samples"}:
            raise ValueError(
                f"report.outlier_ratio_stats[{stat.id}].numerator must be one of new_outlier_samples or outlier_samples."
            )
        if stat.group_by_dimension not in seen_dimensions:
            raise ValueError(
                f"report.outlier_ratio_stats[{stat.id}].group_by_dimension must match one of the row_dimensions."
            )

    for index, node_order in enumerate(_effective_node_orders(template.report), start=1):
        if not node_order:
            raise ValueError(f"report.node_orders[{index}] cannot be empty.")
        if len(set(node_order)) != len(node_order):
            raise ValueError(
                f"report.node_orders[{index}] contains duplicate nodes."
            )
    _validate_node_order_consistency(_effective_node_orders(template.report))

    if template.golden_reference_defaults.center_method not in {"mean", "median"}:
        raise ValueError(
            "golden_reference_defaults.center_method must be one of mean or median."
        )
    if template.golden_reference_defaults.threshold_mode not in {
        "relative",
        "sigma",
        "hybrid",
    }:
        raise ValueError(
            "golden_reference_defaults.threshold_mode must be one of relative, sigma, or hybrid."
        )


def _load_template_workbook_payload(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading template xlsx requires openpyxl. Run `pip3 install -e .` first."
        ) from exc

    workbook = load_workbook(path, data_only=False)
    info = _load_template_info_sheet(workbook)
    payload: dict[str, Any] = {
        "name": info["name"],
        "header_row": info["header_row"],
        "measurement_header_row": info["measurement_header_row"],
        "row_header_row": info["row_header_row"],
        "unit_row": info["unit_row"],
        "data_start_row": info["data_start_row"],
        "row_dimensions": _load_row_dimensions_sheet(workbook),
        "analysis_groups": _load_analysis_groups_sheet(workbook),
            "golden_reference_defaults": {
                "reference_dimensions": info["golden_reference_dimensions"],
                "filters": info["golden_reference_filters"],
                "center_method": info["golden_center_method"],
                "threshold_mode": info["golden_threshold_mode"],
                "relative_limit": info["golden_relative_limit"],
                "sigma_multiplier": info["golden_sigma_multiplier"],
            },
        "report": {
            "golden_values": _load_key_value_sheet(
                workbook,
                TEMPLATE_WORKBOOK_SHEETS["golden_values"],
                key_header="key",
                value_header="value",
                numeric_values=True,
            ),
            "zscore_thresholds": {"default": info["report_default_zscore_threshold"]},
            "golden_deviation_thresholds": {"default": info["golden_relative_limit"]},
            "node_orders": _load_node_orders_sheet(workbook, fallback=info["node_order"]),
            "outlier_ratio_stats": _load_outlier_ratio_stats_sheet(workbook),
            "outlier_fail_method": info["outlier_fail_method"],
            "outlier_chain_fail_rule": info["outlier_chain_fail_rule"],
        },
    }
    return payload


def _save_template_workbook(payload: dict[str, Any], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "Writing template xlsx requires openpyxl. Run `pip3 install -e .` first."
        ) from exc

    workbook = Workbook()
    info_sheet = workbook.active
    info_sheet.title = TEMPLATE_WORKBOOK_SHEETS["info"]
    info_rows = [
        ("field", "value", "notes"),
        ("name", payload["name"], "Template name"),
        ("header_row", payload.get("header_row", 0), "Zero-based"),
        (
            "measurement_header_row",
            payload.get("measurement_header_row"),
            "Zero-based; leave blank to follow header_row",
        ),
        (
            "row_header_row",
            payload.get("row_header_row"),
            "Zero-based; leave blank to follow measurement_header_row",
        ),
        ("unit_row", payload.get("unit_row"), "Zero-based; optional"),
        ("data_start_row", payload.get("data_start_row"), "Zero-based; optional"),
        (
            "node_order",
            None,
            "Deprecated in Excel template; edit node_orders sheet instead",
        ),
        (
            "outlier_fail_method",
            payload.get("report", {}).get("outlier_fail_method", "modified_z_score"),
            "modified_z_score / golden_deviation / zscore_and_golden / zscore_or_golden",
        ),
        (
            "outlier_chain_fail_rule",
            payload.get("report", {}).get("outlier_chain_fail_rule", "any_fail"),
            "any_fail or all_fail",
        ),
        (
            "report_default_zscore_threshold",
            payload.get("report", {}).get("zscore_thresholds", {}).get("default", 3.5),
            "Default zscore threshold used by report/export",
        ),
        (
            "golden_reference_dimensions",
            ",".join(
                payload.get("golden_reference_defaults", {}).get(
                    "reference_dimensions", []
                )
            ),
            "Comma-separated default reference dims for Build Golden",
        ),
        (
            "golden_reference_filters",
            _format_filters_cell(
                payload.get("golden_reference_defaults", {}).get("filters", {})
            ),
            "Semicolon-separated key=value defaults for Build Golden",
        ),
        (
            "golden_reference_center_method",
            payload.get("golden_reference_defaults", {}).get("center_method", "mean"),
            "mean or median",
        ),
        (
            "golden_reference_threshold_mode",
            payload.get("golden_reference_defaults", {}).get("threshold_mode", "relative"),
            "relative / sigma / hybrid",
        ),
        (
            "golden_reference_relative_limit",
            payload.get("golden_reference_defaults", {}).get("relative_limit", 0.2),
            "Optional numeric default for relative/hybrid",
        ),
        (
            "golden_reference_sigma_multiplier",
            payload.get("golden_reference_defaults", {}).get("sigma_multiplier", 3.0),
            "Optional numeric default for sigma/hybrid",
        ),
    ]
    _write_sheet_rows(info_sheet, info_rows, wrap_columns={2, 3})
    _style_header_row(info_sheet, fill_color="1F4E78", font_color="FFFFFF")
    info_sheet.freeze_panes = "A2"
    info_sheet.column_dimensions["A"].width = 28
    info_sheet.column_dimensions["B"].width = 32
    info_sheet.column_dimensions["C"].width = 42

    row_dimensions_sheet = workbook.create_sheet(TEMPLATE_WORKBOOK_SHEETS["row_dimensions"])
    max_source_delimiters = max(
        1,
        max(
            (
                len(source.get("split_delimiters", []))
                for item in payload.get("row_dimensions", [])
                for source in item.get("sources", [])
            ),
            default=0,
        ),
    )
    row_dimensions_rows = [
        tuple(
            [
                "name",
                "display_name",
                "optional",
                "default_value",
                "source_order",
                "source_column",
                "split_position",
                "skip_empty_parts",
                *_indexed_headers("delimiter", max_source_delimiters),
            ]
        )
    ]
    for item in payload.get("row_dimensions", []):
        sources = item.get("sources", []) or [{}]
        for source_index, source in enumerate(sources, start=1):
            row_dimensions_rows.append(
                tuple(
                    [
                        item["name"],
                        item.get("display_name"),
                        item.get("optional", False),
                        item.get("default_value"),
                        source_index,
                        source.get("column"),
                        source.get("split_position"),
                        source.get("skip_empty_parts", False),
                        *_pad_list(source.get("split_delimiters", []), max_source_delimiters),
                    ]
                )
            )
    _write_sheet_rows(row_dimensions_sheet, row_dimensions_rows)
    _style_header_row(row_dimensions_sheet, fill_color="2F75B5", font_color="FFFFFF")
    row_dimensions_sheet.freeze_panes = "A2"
    _set_sheet_widths(
        row_dimensions_sheet,
        _merge_width_maps(
            {
                "A": 24,
                "B": 24,
                "C": 12,
                "D": 18,
                "E": 14,
                "F": 28,
                "G": 14,
                "H": 16,
            },
            _indexed_widths(start_column=9, count=max_source_delimiters, width=14),
        ),
    )

    analysis_groups_sheet = workbook.create_sheet(TEMPLATE_WORKBOOK_SHEETS["analysis_groups"])
    analysis_group_rows = [
        (
            "id",
            "display_name",
            "analysis_mode",
            "presentation_group",
            "unit",
            "column_order",
            "column_name",
        )
    ]
    for item in payload.get("analysis_groups", []):
        group_columns = item.get("columns", []) or [None]
        for index, column_name in enumerate(group_columns, start=1):
            analysis_group_rows.append(
                (
                    item["id"],
                    item.get("display_name"),
                    item["analysis_mode"],
                    item.get("presentation_group"),
                    item.get("unit"),
                    index,
                    column_name,
                )
            )
    _write_sheet_rows(analysis_groups_sheet, analysis_group_rows)
    _style_header_row(analysis_groups_sheet, fill_color="5B9BD5", font_color="FFFFFF")
    analysis_groups_sheet.freeze_panes = "A2"
    _set_sheet_widths(
        analysis_groups_sheet,
        {
            "A": 24,
            "B": 24,
            "C": 18,
            "D": 20,
            "E": 12,
            "F": 14,
            "G": 36,
        },
    )

    node_orders_sheet = workbook.create_sheet(TEMPLATE_WORKBOOK_SHEETS["node_orders"])
    node_orders = _payload_node_orders(payload.get("report", {}))
    max_nodes = max(1, max((len(item) for item in node_orders), default=0))
    node_order_rows = [tuple(["sequence_name", *_indexed_headers("node", max_nodes)])]
    for index, node_order in enumerate(node_orders, start=1):
        node_order_rows.append(
            tuple([f"Sequence {index}", *_pad_list(node_order, max_nodes)])
        )
    if len(node_order_rows) == 1:
        node_order_rows.append(tuple(["Sequence 1", *_pad_list([], max_nodes)]))
    _write_sheet_rows(node_orders_sheet, node_order_rows)
    _style_header_row(node_orders_sheet, fill_color="8064A2", font_color="FFFFFF")
    node_orders_sheet.freeze_panes = "A2"
    _set_sheet_widths(
        node_orders_sheet,
        _merge_width_maps(
            {"A": 20},
            _indexed_widths(start_column=2, count=max_nodes, width=18),
        ),
    )

    golden_values_sheet = workbook.create_sheet(TEMPLATE_WORKBOOK_SHEETS["golden_values"])
    golden_rows = [("key", "value")]
    for key, value in payload.get("report", {}).get("golden_values", {}).items():
        golden_rows.append((key, value))
    _write_sheet_rows(golden_values_sheet, golden_rows)
    _style_header_row(golden_values_sheet, fill_color="70AD47", font_color="FFFFFF")
    golden_values_sheet.freeze_panes = "A2"
    _set_sheet_widths(golden_values_sheet, {"A": 40, "B": 16})

    ratio_stats_sheet = workbook.create_sheet(TEMPLATE_WORKBOOK_SHEETS["outlier_ratio_stats"])
    ratio_stat_rows = [
        (
            "id",
            "display_name",
            "group_by_dimension",
            "numerator",
            "filter_type",
            "filter_order",
            "filter_value",
        )
    ]
    for item in payload.get("report", {}).get("outlier_ratio_stats", []):
        filters: list[tuple[str, str]] = []
        filters.extend(("chain", value) for value in item.get("chains", []))
        filters.extend(("raw_column", value) for value in item.get("raw_columns", []))
        filters.extend(("logical_metric", value) for value in item.get("logical_metrics", []))
        if not filters:
            filters = [("", "")]
        for index, (filter_type, filter_value) in enumerate(filters, start=1):
            ratio_stat_rows.append(
                (
                    item["id"],
                    item.get("display_name"),
                    item.get("group_by_dimension", "reliability_node"),
                    item.get("numerator", "new_outlier_samples"),
                    filter_type,
                    index,
                    filter_value,
                )
            )
    if len(ratio_stat_rows) == 1:
        ratio_stat_rows.append(
            ("ratio_by_node_new", "New Outlier Ratio By Node", "reliability_node", "new_outlier_samples", "", 1, "")
        )
    _write_sheet_rows(ratio_stats_sheet, ratio_stat_rows)
    _style_header_row(ratio_stats_sheet, fill_color="A5A5A5", font_color="FFFFFF")
    ratio_stats_sheet.freeze_panes = "A2"
    _set_sheet_widths(
        ratio_stats_sheet,
        {"A": 24, "B": 28, "C": 22, "D": 24, "E": 16, "F": 14, "G": 36},
    )

    readme_sheet = workbook.create_sheet(TEMPLATE_WORKBOOK_SHEETS["readme"])
    readme_rows = [
        ("section", "content"),
        ("Purpose", "This workbook is an editable view of the JSON template."),
        ("Rows", "Every sheet uses the first row as field names."),
        ("Lists", "Excel template writes one item per cell. Older newline-based files are still accepted."),
        ("Row Indexes", "header_row and related row indexes are zero-based."),
        (
            "Node Orders",
            "Use node_orders sheet for multiple node sequences; one node per cell.",
        ),
        (
            "Thresholds",
            "Default zscore threshold and golden relative limit are both edited in template_info.",
        ),
        (
            "Golden Keys",
            "golden_values keys can be group_id or group_id::raw_column.",
        ),
        (
            "Outlier Ratio Stats",
            "Use outlier_ratio_stats sheet to define ratio rows for Outliner Summary. numerator supports new_outlier_samples or outlier_samples. filter_type supports chain, raw_column, logical_metric.",
        ),
    ]
    _write_sheet_rows(readme_sheet, readme_rows, wrap_columns={2})
    _style_header_row(readme_sheet, fill_color="7F7F7F", font_color="FFFFFF")
    readme_sheet.freeze_panes = "A2"
    _set_sheet_widths(readme_sheet, {"A": 18, "B": 80})
    for cell in readme_sheet[1]:
        cell.font = Font(bold=True)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row > 1 and isinstance(cell.value, str) and "\n" in cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)


def _load_template_info_sheet(workbook) -> dict[str, Any]:
    sheet_name = TEMPLATE_WORKBOOK_SHEETS["info"]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Template workbook is missing required sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet {sheet_name} is empty.")
    header = [_normalize_header(item) for item in rows[0]]
    if "field" not in header or "value" not in header:
        raise ValueError(
            f"Sheet {sheet_name} must contain 'field' and 'value' headers."
        )
    field_index = header.index("field")
    value_index = header.index("value")
    values: dict[str, Any] = {}
    for row in rows[1:]:
        field = _clean_cell(row[field_index] if field_index < len(row) else None)
        if not field:
            continue
        values[field] = row[value_index] if value_index < len(row) else None

    return {
        "name": _require_string(values, "name", sheet_name),
        "header_row": _parse_int_value(values.get("header_row"), field_name="header_row"),
        "measurement_header_row": _parse_optional_int_value(
            values.get("measurement_header_row")
        ),
        "row_header_row": _parse_optional_int_value(values.get("row_header_row")),
        "unit_row": _parse_optional_int_value(values.get("unit_row")),
        "data_start_row": _parse_optional_int_value(values.get("data_start_row")),
        "node_order": _parse_list_cell(values.get("node_order")),
        "report_default_zscore_threshold": _coalesce_optional_float(
            values.get("report_default_zscore_threshold"),
            3.5,
        ),
        "outlier_fail_method": _parse_optional_string(
            values.get("outlier_fail_method"), default="modified_z_score"
        ),
        "outlier_chain_fail_rule": _parse_optional_string(
            values.get("outlier_chain_fail_rule"), default="any_fail"
        ),
        "golden_reference_dimensions": _parse_csv_list(
            values.get("golden_reference_dimensions")
        ),
        "golden_reference_filters": _parse_filter_cell(
            values.get("golden_reference_filters")
        ),
        "golden_center_method": _parse_optional_string(
            values.get("golden_reference_center_method"),
            default="mean"
        ),
        "golden_threshold_mode": _parse_optional_string(
            values.get("golden_reference_threshold_mode"),
            default="relative"
        ),
        "golden_relative_limit": _parse_optional_float_value(
            values.get("golden_reference_relative_limit")
        ),
        "golden_sigma_multiplier": _parse_optional_float_value(
            values.get("golden_reference_sigma_multiplier")
        ),
    }


def _load_row_dimensions_sheet(workbook) -> list[dict[str, Any]]:
    rows = _load_rows_from_sheet(
        workbook,
        TEMPLATE_WORKBOOK_SHEETS["row_dimensions"],
        required_headers={
            "name",
            "display_name",
            "optional",
            "default_value",
            "source_order",
            "source_column",
            "split_position",
            "skip_empty_parts",
        },
    )
    grouped_rows: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        name = _clean_cell(row.get("name"))
        if not name:
            continue
        grouped_rows.setdefault(name, []).append(row)

    payload: list[dict[str, Any]] = []
    for name, dimension_rows in grouped_rows.items():
        sorted_rows = sorted(
            dimension_rows,
            key=lambda item: (
                _parse_optional_int_value(item.get("source_order")) or 10**9,
                _natural_text_sort_key(_clean_cell(item.get("source_column"))),
            ),
        )
        item: dict[str, Any] = {
            "name": name,
            "sources": [],
        }
        display_name = _first_non_empty(dimension_rows, "display_name")
        if display_name:
            item["display_name"] = display_name
        if _coalesce_bool(dimension_rows, "optional"):
            item["optional"] = True
        default_value = _first_non_empty(dimension_rows, "default_value")
        if default_value:
            item["default_value"] = default_value

        for row in sorted_rows:
            source_column = _clean_cell(row.get("source_column"))
            if not source_column:
                continue
            source_payload: dict[str, Any] = {"column": source_column}
            split_position = _parse_optional_int_value(row.get("split_position"))
            if split_position is not None:
                source_payload["split_position"] = split_position
            if _parse_bool_value(row.get("skip_empty_parts")):
                source_payload["skip_empty_parts"] = True
            split_delimiters = _extract_list_values(
                row,
                single_cell_key="split_delimiters",
                indexed_prefix="delimiter_",
            )
            if split_delimiters:
                source_payload["split_delimiters"] = split_delimiters
            item["sources"].append(source_payload)

        if not item["sources"]:
            raise ValueError(f"Row dimension {name} must define at least one source.")
        payload.append(item)
    return payload


def _load_analysis_groups_sheet(workbook) -> list[dict[str, Any]]:
    rows = _load_rows_from_sheet(
        workbook,
        TEMPLATE_WORKBOOK_SHEETS["analysis_groups"],
        required_headers={"id"},
    )
    if rows and "column_name" in rows[0]:
        return _load_analysis_groups_long_rows(rows)
    payload: list[dict[str, Any]] = []
    for row in rows:
        group_id = _clean_cell(row.get("id"))
        if not group_id:
            continue
        columns = _extract_list_values(row, single_cell_key="columns", indexed_prefix="column_")
        if not columns:
            raise ValueError(f"Analysis group {group_id} must define at least one column.")
        item: dict[str, Any] = {
            "id": group_id,
            "display_name": _clean_cell(row.get("display_name")) or group_id,
            "columns": columns,
            "analysis_mode": _require_clean_value(
                row.get("analysis_mode"),
                field_name=f"{group_id}.analysis_mode",
            ),
        }
        presentation_group = _clean_cell(row.get("presentation_group"))
        if presentation_group:
            item["presentation_group"] = presentation_group
        unit = _clean_cell(row.get("unit"))
        if unit:
            item["unit"] = unit
        payload.append(item)
    return payload


def _load_analysis_groups_long_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped_rows: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        group_id = _clean_cell(row.get("id"))
        if not group_id:
            continue
        grouped_rows.setdefault(group_id, []).append(row)

    payload: list[dict[str, Any]] = []
    for group_id, group_rows in grouped_rows.items():
        sorted_rows = sorted(
            group_rows,
            key=lambda item: (
                _parse_optional_int_value(item.get("column_order")) or 10**9,
                _natural_text_sort_key(_clean_cell(item.get("column_name"))),
            ),
        )
        display_name = _first_non_empty(group_rows, "display_name") or group_id
        analysis_mode = _require_clean_value(
            _first_non_empty(group_rows, "analysis_mode"),
            field_name=f"{group_id}.analysis_mode",
        )
        presentation_group = _first_non_empty(group_rows, "presentation_group")
        unit = _first_non_empty(group_rows, "unit")
        columns = [
            _clean_cell(row.get("column_name"))
            for row in sorted_rows
            if _clean_cell(row.get("column_name"))
        ]
        if not columns:
            raise ValueError(f"Analysis group {group_id} must define at least one column.")
        item: dict[str, Any] = {
            "id": group_id,
            "display_name": display_name,
            "columns": columns,
            "analysis_mode": analysis_mode,
        }
        if presentation_group:
            item["presentation_group"] = presentation_group
        if unit:
            item["unit"] = unit
        payload.append(item)
    return payload


def _load_node_orders_sheet(workbook, fallback: list[str]) -> list[list[str]]:
    sheet_name = TEMPLATE_WORKBOOK_SHEETS["node_orders"]
    if sheet_name not in workbook.sheetnames:
        return [fallback] if fallback else []
    rows = _load_rows_from_sheet(
        workbook,
        sheet_name,
        required_headers={"sequence_name"},
    )
    payload: list[list[str]] = []
    for row in rows:
        nodes = _extract_list_values(row, single_cell_key="nodes", indexed_prefix="node_")
        if nodes:
            payload.append(nodes)
    return payload or ([fallback] if fallback else [])


def _load_outlier_ratio_stats_sheet(workbook) -> list[dict[str, Any]]:
    sheet_name = TEMPLATE_WORKBOOK_SHEETS["outlier_ratio_stats"]
    if sheet_name not in workbook.sheetnames:
        return []
    rows = _load_rows_from_sheet(
        workbook,
        sheet_name,
        required_headers={"id", "group_by_dimension", "numerator"},
    )
    grouped_rows: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        stat_id = _clean_cell(row.get("id"))
        if not stat_id:
            continue
        grouped_rows.setdefault(stat_id, []).append(row)

    payload: list[dict[str, Any]] = []
    for stat_id, stat_rows in grouped_rows.items():
        sorted_rows = sorted(
            stat_rows,
            key=lambda item: (
                _parse_optional_int_value(item.get("filter_order")) or 10**9,
                _natural_text_sort_key(_clean_cell(item.get("filter_value"))),
            ),
        )
        item: dict[str, Any] = {
            "id": stat_id,
            "display_name": _first_non_empty(stat_rows, "display_name") or stat_id,
            "group_by_dimension": _first_non_empty(stat_rows, "group_by_dimension") or "reliability_node",
            "numerator": _first_non_empty(stat_rows, "numerator") or "new_outlier_samples",
            "chains": [],
            "raw_columns": [],
            "logical_metrics": [],
        }
        for row in sorted_rows:
            filter_type = _clean_cell(row.get("filter_type"))
            filter_value = _clean_cell(row.get("filter_value"))
            if not filter_type or not filter_value:
                continue
            if filter_type == "chain":
                item["chains"].append(filter_value)
            elif filter_type == "raw_column":
                item["raw_columns"].append(filter_value)
            elif filter_type == "logical_metric":
                item["logical_metrics"].append(filter_value)
            else:
                raise ValueError(
                    f"Unsupported filter_type in {sheet_name}: {filter_type}. Expected chain, raw_column, or logical_metric."
                )
        payload.append(item)
    return payload


def _load_key_value_sheet(
    workbook,
    sheet_name: str,
    key_header: str,
    value_header: str,
    numeric_values: bool,
) -> dict[str, Any]:
    if sheet_name not in workbook.sheetnames:
        return {}
    rows = _load_rows_from_sheet(
        workbook,
        sheet_name,
        required_headers={key_header, value_header},
    )
    payload: dict[str, Any] = {}
    for row in rows:
        key = _clean_cell(row.get(key_header))
        if not key:
            continue
        raw_value = row.get(value_header)
        if numeric_values:
            payload[key] = _parse_float_value(raw_value, field_name=f"{sheet_name}.{key}")
        else:
            payload[key] = _clean_cell(raw_value)
    return payload


def _load_threshold_sheet(
    workbook,
    sheet_name: str,
    default_value: float,
) -> dict[str, Any]:
    if sheet_name not in workbook.sheetnames:
        return {"default": default_value, "overrides": {}}
    rows = _load_rows_from_sheet(
        workbook,
        sheet_name,
        required_headers={"metric_key", "value"},
    )
    payload: dict[str, Any] = {"default": default_value, "overrides": {}}
    for row in rows:
        metric_key = _clean_cell(row.get("metric_key"))
        if not metric_key:
            continue
        value = _parse_float_value(row.get("value"), field_name=f"{sheet_name}.{metric_key}")
        if metric_key == "default":
            payload["default"] = value
        else:
            payload["overrides"][metric_key] = value
    return payload


def _load_rows_from_sheet(
    workbook,
    sheet_name: str,
    required_headers: set[str],
) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Template workbook is missing required sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet {sheet_name} is empty.")
    header = [_normalize_header(item) for item in rows[0]]
    missing = [item for item in required_headers if item not in header]
    if missing:
        raise ValueError(
            f"Sheet {sheet_name} is missing required columns: {', '.join(sorted(missing))}"
        )
    payload_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        row_payload: dict[str, Any] = {}
        for index, column_name in enumerate(header):
            if not column_name:
                continue
            row_payload[column_name] = row[index] if index < len(row) else None
        if any(_clean_cell(value) not in {"", None} for value in row_payload.values()):
            payload_rows.append(row_payload)
    return payload_rows


def _write_sheet_rows(sheet, rows: list[tuple[Any, ...]], wrap_columns: set[int] | None = None) -> None:
    wrap_columns = wrap_columns or set()
    try:
        from openpyxl.styles import Alignment
    except ImportError:
        Alignment = None
    for row_index, row_values in enumerate(rows, start=1):
        for column_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if column_index in wrap_columns and Alignment is not None:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _style_header_row(sheet, fill_color: str, font_color: str) -> None:
    try:
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    font = Font(color=font_color, bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font


def _set_sheet_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _normalize_header(value: Any) -> str:
    return _clean_cell(value).lower()


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_list_cell(value: Any) -> list[str]:
    text = _clean_cell(value)
    if not text:
        return []
    normalized = text.replace("\r\n", "\n").replace(";", "\n").replace(",", "\n")
    return [item.strip() for item in normalized.split("\n") if item.strip()]


def _parse_csv_list(value: Any) -> list[str]:
    text = _clean_cell(value)
    if not text:
        return []
    return [item.strip() for item in text.split(",") if item.strip()]


def _format_filters_cell(filters: dict[str, str]) -> str:
    return "; ".join(
        f"{key}={value}" for key, value in filters.items() if str(key).strip()
    )


def _parse_filter_cell(value: Any) -> dict[str, str]:
    text = _clean_cell(value)
    if not text:
        return {}
    parts = re.split(r"[;\n\r]+", text)
    filters: dict[str, str] = {}
    for item in parts:
        entry = item.strip()
        if not entry:
            continue
        if "=" not in entry:
            raise ValueError(
                "golden_reference_filters must use key=value pairs separated by ';'."
            )
        key, raw_value = entry.split("=", 1)
        filters[key.strip()] = raw_value.strip()
    return filters


def _parse_bool_value(value: Any) -> bool:
    text = _clean_cell(value).lower()
    if not text:
        return False
    if text in {"1", "true", "yes", "y"}:
        return True
    if text in {"0", "false", "no", "n"}:
        return False
    raise ValueError(f"Invalid boolean value: {value}")


def _parse_optional_string(value: Any, default: str) -> str:
    text = _clean_cell(value)
    return text or default


def _parse_int_value(value: Any, field_name: str) -> int:
    if value is None or _clean_cell(value) == "":
        raise ValueError(f"{field_name} is required.")
    return int(float(value))


def _parse_optional_int_value(value: Any) -> int | None:
    if value is None or _clean_cell(value) == "":
        return None
    return int(float(value))


def _parse_float_value(value: Any, field_name: str) -> float:
    if value is None or _clean_cell(value) == "":
        raise ValueError(f"{field_name} is required.")
    return float(value)


def _parse_optional_float_value(value: Any) -> float | None:
    if value is None or _clean_cell(value) == "":
        return None
    return float(value)


def _coalesce_optional_float(value: Any, default: float) -> float:
    parsed = _parse_optional_float_value(value)
    if parsed is None:
        return default
    return parsed


def _require_string(values: dict[str, Any], field_name: str, sheet_name: str) -> str:
    value = _clean_cell(values.get(field_name))
    if not value:
        raise ValueError(f"Sheet {sheet_name} requires a value for {field_name}.")
    return value


def _require_clean_value(value: Any, field_name: str) -> str:
    text = _clean_cell(value)
    if not text:
        raise ValueError(f"{field_name} is required.")
    return text


def _serialize_number(value: float) -> int | float:
    numeric = float(value)
    if numeric.is_integer():
        return int(numeric)
    return numeric


def _effective_node_orders(report: ReportConfig) -> list[list[str]]:
    if report.node_orders:
        return [list(item) for item in report.node_orders if item]
    if report.node_order:
        return [list(report.node_order)]
    return []


def _payload_node_orders(report_payload: dict[str, Any]) -> list[list[str]]:
    if "node_orders" in report_payload:
        return [
            [str(node).strip() for node in item if str(node).strip()]
            for item in report_payload.get("node_orders", [])
            if item
        ]
    primary = _primary_node_order(report_payload)
    return [primary] if primary else []


def _primary_node_order(report_payload: dict[str, Any]) -> list[str]:
    if "node_orders" in report_payload and report_payload.get("node_orders"):
        first = report_payload["node_orders"][0]
        return [str(node).strip() for node in first if str(node).strip()]
    return [str(node).strip() for node in report_payload.get("node_order", []) if str(node).strip()]


def _validate_node_order_consistency(node_orders: list[list[str]]) -> None:
    seen_pairs: dict[tuple[str, str], int] = {}
    for sequence_index, node_order in enumerate(node_orders, start=1):
        for left_index, left_node in enumerate(node_order):
            for right_node in node_order[left_index + 1 :]:
                reverse_key = (right_node, left_node)
                if reverse_key in seen_pairs:
                    other_sequence_index = seen_pairs[reverse_key]
                    raise ValueError(
                        "report.node_orders contains conflicting node directions: "
                        f"sequence {other_sequence_index} requires {right_node} before {left_node}, "
                        f"but sequence {sequence_index} requires {left_node} before {right_node}."
                    )
                seen_pairs[(left_node, right_node)] = sequence_index


def _indexed_headers(prefix: str, count: int) -> list[str]:
    return [f"{prefix}_{index}" for index in range(1, count + 1)]


def _pad_list(values: list[Any], count: int) -> list[Any]:
    padded = list(values[:count])
    while len(padded) < count:
        padded.append(None)
    return padded


def _merge_width_maps(*maps: dict[str, float]) -> dict[str, float]:
    merged: dict[str, float] = {}
    for item in maps:
        merged.update(item)
    return merged


def _indexed_widths(start_column: int, count: int, width: float) -> dict[str, float]:
    return {
        _column_letter_from_index(start_column + offset): width
        for offset in range(count)
    }


def _column_letter_from_index(index: int) -> str:
    letters: list[str] = []
    value = index
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def _extract_list_values(
    row: dict[str, Any],
    single_cell_key: str,
    indexed_prefix: str,
) -> list[str]:
    indexed_values = [
        (key, value)
        for key, value in row.items()
        if key.startswith(indexed_prefix)
    ]
    if indexed_values:
        return [
            _clean_cell(value)
            for key, value in sorted(indexed_values, key=lambda item: _indexed_key_sort(item[0], indexed_prefix))
            if _clean_cell(value)
        ]
    return _parse_list_cell(row.get(single_cell_key))


def _indexed_key_sort(key: str, prefix: str) -> int:
    suffix = key[len(prefix) :]
    if suffix.isdigit():
        return int(suffix)
    return 10**9


def _first_non_empty(rows: list[dict[str, Any]], key: str) -> str:
    for row in rows:
        value = _clean_cell(row.get(key))
        if value:
            return value
    return ""


def _coalesce_bool(rows: list[dict[str, Any]], key: str) -> bool:
    for row in rows:
        if _parse_bool_value(row.get(key)):
            return True
    return False


def _natural_text_sort_key(value: str) -> tuple[Any, ...]:
    return _split_natural_parts(value)


def _split_natural_parts(value: str) -> tuple[Any, ...]:
    parts = re.split(r"(\d+)", value or "")
    key: list[Any] = []
    for part in parts:
        if not part:
            continue
        if part.isdigit():
            key.append(int(part))
        else:
            key.append(part.lower())
    return tuple(key)
