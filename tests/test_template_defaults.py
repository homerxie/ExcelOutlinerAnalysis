from __future__ import annotations

from pathlib import Path
import json
import tempfile
import unittest

from excel_data_analysis.template import load_template, save_template


class TemplateDefaultsTests(unittest.TestCase):
    def test_golden_reference_defaults_roundtrip_in_json_template(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            template_path = Path(temp_dir) / "template.json"
            template_path.write_text(
                """
{
  "name": "demo",
  "row_dimensions": [],
  "analysis_groups": [],
  "golden_reference_defaults": {
    "reference_dimensions": ["sample_id"],
    "filters": {
      "reliability_node": "T0"
    },
    "center_method": "median",
    "threshold_mode": "hybrid",
    "relative_limit": 0.15,
    "sigma_multiplier": 2.5
  },
  "report": {
    "zscore_thresholds": {
      "default": 3.5
    },
    "golden_deviation_thresholds": {
      "default": 0.05
    },
    "outlier_fail_method": "zscore_or_golden",
    "outlier_chain_fail_rule": "any_fail",
    "node_order": ["T0", "T1"]
  }
}
""".strip(),
                encoding="utf-8",
            )

            template = load_template(template_path)
            self.assertEqual(
                template.golden_reference_defaults.reference_dimensions,
                ["sample_id"],
            )
            self.assertEqual(
                template.golden_reference_defaults.filters,
                {"reliability_node": "T0"},
            )
            self.assertEqual(template.golden_reference_defaults.center_method, "median")
            self.assertEqual(template.golden_reference_defaults.threshold_mode, "hybrid")
            self.assertEqual(template.golden_reference_defaults.relative_limit, 0.15)
            self.assertEqual(template.golden_reference_defaults.sigma_multiplier, 2.5)

            output_path = Path(temp_dir) / "template_roundtrip.json"
            save_template(template, output_path, if_exists="overwrite")
            saved_payload = json.loads(output_path.read_text(encoding="utf-8"))
            roundtrip = load_template(output_path)

            self.assertEqual(
                roundtrip.golden_reference_defaults.reference_dimensions,
                ["sample_id"],
            )
            self.assertEqual(
                roundtrip.golden_reference_defaults.filters,
                {"reliability_node": "T0"},
            )
            self.assertEqual(roundtrip.golden_reference_defaults.center_method, "median")
            self.assertEqual(roundtrip.golden_reference_defaults.threshold_mode, "hybrid")
            self.assertEqual(roundtrip.golden_reference_defaults.relative_limit, 0.15)
            self.assertEqual(roundtrip.golden_reference_defaults.sigma_multiplier, 2.5)
            self.assertNotIn("golden_deviation_thresholds", saved_payload.get("report", {}))

    def test_template_xlsx_writes_threshold_defaults_to_info_sheet_only(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            template_path = Path(temp_dir) / "template.json"
            template_path.write_text(
                """
{
  "name": "demo",
  "row_dimensions": [],
  "analysis_groups": [],
  "golden_reference_defaults": {
    "reference_dimensions": ["sample_id"],
    "filters": {
      "reliability_node": "T0"
    },
    "center_method": "mean",
    "threshold_mode": "relative",
    "relative_limit": 0.2,
    "sigma_multiplier": 3.0
  },
  "report": {
    "zscore_thresholds": {
      "default": 4.2,
      "overrides": {
        "Link4": 5.1
      }
    },
    "golden_deviation_thresholds": {
      "default": 0.08,
      "overrides": {
        "Link4": 0.12
      }
    },
    "outlier_fail_method": "golden_deviation",
    "outlier_chain_fail_rule": "any_fail",
    "node_order": ["T0", "T1"]
  }
}
""".strip(),
                encoding="utf-8",
            )
            template = load_template(template_path)
            workbook_path = Path(temp_dir) / "template.xlsx"
            save_template(template, workbook_path, if_exists="overwrite")

            from openpyxl import load_workbook

            workbook = load_workbook(workbook_path, data_only=False)
            info_sheet = workbook["template_info"]
            info_values = {
                str(row[0].value): row[1].value
                for row in info_sheet.iter_rows(min_row=2, max_col=2)
                if row[0].value
            }
            self.assertEqual(info_values["report_default_zscore_threshold"], 4.2)
            self.assertEqual(info_values["golden_reference_relative_limit"], 0.2)
            self.assertNotIn("zscore_thresholds", workbook.sheetnames)
            self.assertNotIn("golden_deviation_thresholds", workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
